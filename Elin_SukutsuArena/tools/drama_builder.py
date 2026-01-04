"""
DramaBuilder - CWL ドラマファイル生成用ビルダークラス

CWL (Custom Whatever Loader) のドラマファイル形式に準拠した
Excel ファイルを生成するためのビルダーパターン実装。

フォーマット仕様は docs/cwl_drama_format.md を参照。
"""

import openpyxl
import os
from typing import Optional, List, Union, Dict, Any, Set

# フラグ定義をインポート（検証用）
try:
    from flag_definitions import (
        get_all_flags, get_all_enums,
        EnumFlag, IntFlag, BoolFlag, StringFlag,
        PREFIX as FLAG_PREFIX
    )
    FLAG_VALIDATION_ENABLED = True
except ImportError:
    FLAG_VALIDATION_ENABLED = False
    print("[DramaBuilder] Warning: flag_definitions not found, flag validation disabled")


def _validate_flag(flag_key: str, value: Any) -> List[str]:
    """
    フラグキーと値を検証。
    エラーがあればメッセージのリストを返す。

    Args:
        flag_key: フラグキー（例: "chitsii.arena.player.motivation"）
        value: 設定する値

    Returns:
        エラーメッセージのリスト（空なら問題なし）
    """
    if not FLAG_VALIDATION_ENABLED:
        return []

    errors = []

    # プレフィクスを持つフラグのみ検証
    if not flag_key.startswith(FLAG_PREFIX + "."):
        # 旧来のフラグ（sukutsu_gladiator等）はスキップ
        return []

    # 登録されたフラグか確認
    all_flags = {f.full_key: f for f in get_all_flags()}

    if flag_key not in all_flags:
        errors.append(f"Unknown flag key: '{flag_key}'")
        return errors

    flag_def = all_flags[flag_key]

    # 型チェック
    if isinstance(flag_def, EnumFlag):
        if flag_def.enum_type:
            valid_ordinals = list(range(len(flag_def.enum_type)))
            if isinstance(value, int) and value not in valid_ordinals and value != -1:
                errors.append(
                    f"Invalid ordinal {value} for enum flag '{flag_key}'. "
                    f"Valid: {valid_ordinals} or -1 (null)"
                )
    elif isinstance(flag_def, IntFlag):
        if not isinstance(value, int):
            errors.append(f"Flag '{flag_key}' expects int, got {type(value).__name__}")
        elif hasattr(flag_def, 'min_value') and hasattr(flag_def, 'max_value'):
            if value < flag_def.min_value or value > flag_def.max_value:
                errors.append(
                    f"Value {value} out of range for '{flag_key}'. "
                    f"Expected: [{flag_def.min_value}, {flag_def.max_value}]"
                )
    elif isinstance(flag_def, BoolFlag):
        if value not in (0, 1, True, False):
            errors.append(f"Flag '{flag_key}' expects bool (0/1), got {value}")

    return errors

# CWL準拠ヘッダー (if2なし、version/text列あり)
HEADERS = ['step', 'jump', 'if', 'action', 'param', 'actor', 'version', 'id', 'text_JP', 'text_EN', 'text']


class DramaLabel:
    """ステップラベルを表すクラス"""
    def __init__(self, key: str):
        self.key = key
    def __str__(self):
        return self.key


class DramaActor:
    """アクターを表すクラス"""
    def __init__(self, key: str):
        self.key = key
    def __str__(self):
        return self.key


class DramaBuilder:
    """
    CWLドラマファイルを構築するビルダークラス。

    使用例:
        drama = DramaBuilder()
        pc = drama.register_actor("pc", "プレイヤー", "Player")
        npc = drama.register_actor("master", "マスター", "Master")

        lbl_main = drama.label("main")
        drama.step(lbl_main)
        drama.say("greet", "こんにちは", actor=npc)
        drama.finish()

        drama.save("drama.xlsx", sheet_name="master")
    """

    def __init__(self) -> None:
        self.entries: List[Dict[str, Any]] = []
        self.actors: Dict[str, Dict[str, str]] = {}
        self.registered_steps: Set[str] = set()
        self._current_step: Optional[str] = None

    def _resolve_key(self, obj: Union[str, DramaLabel, DramaActor]) -> str:
        """オブジェクトからキーを取得"""
        if isinstance(obj, (DramaLabel, DramaActor)):
            return obj.key
        return obj

    def register_actor(self, actor_id: str, name_jp: str, name_en: str = "") -> DramaActor:
        """アクターを登録"""
        self.actors[actor_id] = {'jp': name_jp, 'en': name_en or name_jp}
        return DramaActor(actor_id)

    def label(self, key: str) -> DramaLabel:
        """ラベルを作成"""
        return DramaLabel(key)

    def step(self, step_label: Union[str, DramaLabel]) -> 'DramaBuilder':
        """
        新しいステップを開始。
        CWL形式では、step行は step列のみを含み、内容は次の行から。
        """
        key = self._resolve_key(step_label)
        self.registered_steps.add(key)
        self._current_step = key
        # step行を追加（step列のみ）
        self.entries.append({'step': key})
        return self

    def say(self, text_id: str, text_jp: str, text_en: str = "",
            actor: Union[str, DramaActor] = None) -> 'DramaBuilder':
        """テキスト行を追加"""
        actor_key = self._resolve_key(actor) if actor else None
        entry = {
            'id': text_id,
            'text_JP': text_jp,
            'text_EN': text_en or text_jp,
        }
        if actor_key:
            entry['actor'] = actor_key
        self.entries.append(entry)
        return self

    def choice(self, jump_to: Union[str, DramaLabel], text_jp: str,
               text_en: str = "", text_id: str = "") -> 'DramaBuilder':
        """選択肢を追加"""
        key = self._resolve_key(jump_to)
        entry = {
            'action': 'choice',
            'jump': key,
            'text_JP': text_jp,
            'text_EN': text_en or text_jp,
        }
        if text_id:
            entry['id'] = text_id
        self.entries.append(entry)
        return self

    def jump(self, jump_to: Union[str, DramaLabel]) -> 'DramaBuilder':
        """指定ステップにジャンプ"""
        key = self._resolve_key(jump_to)
        self.entries.append({'jump': key})
        return self

    def branch_if(self, flag: str, operator: str, value: int,
                  jump_to: Union[str, DramaLabel],
                  actor: Union[str, DramaActor] = None) -> 'DramaBuilder':
        """
        条件分岐。invoke* + if_flag を使用。
        """
        key = self._resolve_key(jump_to)
        actor_key = self._resolve_key(actor) if actor else 'pc'
        entry = {
            'action': 'invoke*',
            'param': f'if_flag({flag}, {operator}{value})',
            'jump': key,
            'actor': actor_key,
        }
        self.entries.append(entry)
        return self

    def set_flag(self, flag: str, value: int = 1) -> 'DramaBuilder':
        """フラグを設定（検証付き）"""
        # フラグ検証
        errors = _validate_flag(flag, value)
        if errors:
            for err in errors:
                self._add_validation_error(f"set_flag: {err}")

        self.entries.append({
            'action': 'setFlag',
            'param': f'{flag},{value}',
        })
        return self

    def mod_flag(self, flag: str, operator: str, value: int,
                 actor: Union[str, DramaActor] = None) -> 'DramaBuilder':
        """フラグを変更 (invoke* mod_flag) - 検証付き"""
        # mod_flag の場合、絶対値で検証（加減算なので範囲チェックは緩め）
        if operator == "=" :
            errors = _validate_flag(flag, value)
            if errors:
                for err in errors:
                    self._add_validation_error(f"mod_flag: {err}")

        actor_key = self._resolve_key(actor) if actor else 'pc'
        self.entries.append({
            'action': 'invoke*',
            'param': f'mod_flag({flag}, {operator}{value})',
            'actor': actor_key,
        })
        return self

    def _add_validation_error(self, message: str) -> None:
        """検証エラーを追加"""
        if not hasattr(self, '_validation_errors'):
            self._validation_errors = []
        self._validation_errors.append(message)
        print(f"[DramaBuilder] ERROR: {message}")

    def get_validation_errors(self) -> List[str]:
        """検証エラーを取得"""
        return getattr(self, '_validation_errors', [])

    def action(self, action_name: str, param: str = None,
               jump: Union[str, DramaLabel] = None,
               actor: Union[str, DramaActor] = None) -> 'DramaBuilder':
        """汎用アクションを追加"""
        entry = {'action': action_name}
        if param:
            entry['param'] = param
        if jump:
            entry['jump'] = self._resolve_key(jump)
        if actor:
            entry['actor'] = self._resolve_key(actor)
        self.entries.append(entry)
        return self

    def on_cancel(self, jump_to: Union[str, DramaLabel]) -> 'DramaBuilder':
        """キャンセル時の動作を設定"""
        key = self._resolve_key(jump_to)
        self.entries.append({'action': 'cancel', 'jump': key})
        return self

    def finish(self) -> 'DramaBuilder':
        """ドラマを終了"""
        self.entries.append({'action': 'end'})
        return self

    def play_sound(self, sound_id: str) -> 'DramaBuilder':
        """効果音を再生"""
        self.entries.append({'action': 'sound', 'param': sound_id})
        return self

    def play_bgm(self, bgm_id: str) -> 'DramaBuilder':
        """BGMを再生"""
        self.entries.append({'action': 'bgm', 'param': bgm_id})
        return self

    def wait(self, seconds: float) -> 'DramaBuilder':
        """待機"""
        self.entries.append({'action': 'wait', 'param': str(seconds)})
        return self

    def effect(self, effect_id: str) -> 'DramaBuilder':
        """エフェクトを再生"""
        self.entries.append({'action': effect_id})
        return self

    # ============================================================================
    # CWL 拡張機能: 組み込みジャンプ
    # ============================================================================

    def inject_unique(self) -> 'DramaBuilder':
        """
        inject/Unique アクションを実行
        標準的な会話UIを挿入（パーティ加入、商品購入など）
        """
        self.entries.append({'action': 'inject', 'param': 'Unique'})
        return self

    def jump_to_trade(self) -> 'DramaBuilder':
        """商人との取引画面にジャンプ (actor需要はtgが対象)"""
        self.entries.append({'jump': '_trade'})
        return self

    def jump_to_buy(self) -> 'DramaBuilder':
        """購入画面にジャンプ"""
        self.entries.append({'jump': '_buy'})
        return self

    def jump_to_join_party(self) -> 'DramaBuilder':
        """パーティ加入にジャンプ"""
        self.entries.append({'jump': '_joinParty'})
        return self

    def jump_to_leave_party(self) -> 'DramaBuilder':
        """パーティ離脱にジャンプ"""
        self.entries.append({'jump': '_leaveParty'})
        return self

    def jump_to_train(self) -> 'DramaBuilder':
        """訓練画面にジャンプ"""
        self.entries.append({'jump': '_train'})
        return self

    def jump_to_heal(self) -> 'DramaBuilder':
        """回復サービスにジャンプ"""
        self.entries.append({'jump': '_heal'})
        return self

    def jump_to_invest_shop(self) -> 'DramaBuilder':
        """店への投資にジャンプ"""
        self.entries.append({'jump': '_investShop'})
        return self

    def jump_to_sell_fame(self) -> 'DramaBuilder':
        """名声売却にジャンプ"""
        self.entries.append({'jump': '_sellFame'})
        return self

    def jump_to_copy_item(self) -> 'DramaBuilder':
        """アイテム複製にジャンプ"""
        self.entries.append({'jump': '_copyItem'})
        return self

    def jump_to_give(self) -> 'DramaBuilder':
        """アイテム渡すにジャンプ"""
        self.entries.append({'jump': '_give'})
        return self

    def jump_to_whore(self) -> 'DramaBuilder':
        """売春サービスにジャンプ (18+)"""
        self.entries.append({'jump': '_whore'})
        return self

    def jump_to_tail(self) -> 'DramaBuilder':
        """尾行にジャンプ"""
        self.entries.append({'jump': '_tail'})
        return self

    def jump_to_suck(self) -> 'DramaBuilder':
        """吸血にジャンプ (サキュバス等)"""
        self.entries.append({'jump': '_suck'})
        return self

    def jump_to_bout(self) -> 'DramaBuilder':
        """決闘にジャンプ"""
        self.entries.append({'jump': '_bout'})
        return self

    def jump_to_rumor(self) -> 'DramaBuilder':
        """噂話にジャンプ"""
        self.entries.append({'jump': '_rumor'})
        return self

    def jump_to_news(self) -> 'DramaBuilder':
        """ニュースにジャンプ"""
        self.entries.append({'jump': '_news'})
        return self

    def jump_builtin(self, builtin_name: str) -> 'DramaBuilder':
        """
        汎用組み込みジャンプ

        Args:
            builtin_name: 組み込みステップ名 (例: "_rumor", "_bout", "_news")
        """
        if not builtin_name.startswith('_'):
            builtin_name = '_' + builtin_name
        self.entries.append({'jump': builtin_name})
        return self

    # ============================================================================
    # CWL 拡張機能: ランタイムスクリプト
    # ============================================================================

    def cs_eval(self, code: str, actor: Union[str, DramaActor] = None) -> 'DramaBuilder':
        """
        C# コードをランタイムで実行 (eval アクション)

        Args:
            code: 実行するC#コード
            actor: 対象アクター (デフォルト: tg)

        Note:
            - Script["var"] で変数を複数 eval 間で共有可能
            - 複雑なロジックは DLL に実装することを推奨
        """
        actor_key = self._resolve_key(actor) if actor else None
        entry = {
            'action': 'eval',
            'param': code,
        }
        if actor_key:
            entry['actor'] = actor_key
        self.entries.append(entry)
        return self

    def cs_script_var_set(self, var_name: str, value_expr: str) -> 'DramaBuilder':
        """
        スクリプト共有変数を設定

        Args:
            var_name: 変数名
            value_expr: C#式として評価される値
        """
        return self.cs_eval(f'Script["{var_name}"] = {value_expr};')

    def cs_script_var_get(self, var_name: str, cast_type: str = "int") -> str:
        """
        スクリプト共有変数を取得するC#式を生成

        Args:
            var_name: 変数名
            cast_type: キャスト型 (例: "int", "string", "bool")

        Returns:
            C#式文字列 (例: "(int)Script[\"counter\"]")
        """
        return f'({cast_type})Script["{var_name}"]'

    # ============================================================================
    # CWL 拡張機能: 条件付きエントリ
    # ============================================================================

    def say_if(self, text_id: str, text_jp: str, condition: str,
               text_en: str = "", actor: Union[str, DramaActor] = None) -> 'DramaBuilder':
        """
        条件付きテキスト行を追加

        Args:
            text_id: テキストID
            text_jp: 日本語テキスト
            condition: if列の条件式 (例: "hasFlag,sukutsu_gladiator" or "=,flag,1")
            text_en: 英語テキスト
            actor: 発言者
        """
        actor_key = self._resolve_key(actor) if actor else None
        entry = {
            'if': condition,
            'id': text_id,
            'text_JP': text_jp,
            'text_EN': text_en or text_jp,
        }
        if actor_key:
            entry['actor'] = actor_key
        self.entries.append(entry)
        return self

    def choice_if(self, jump_to: Union[str, DramaLabel], text_jp: str,
                  condition: str, text_en: str = "", text_id: str = "") -> 'DramaBuilder':
        """
        条件付き選択肢を追加

        Args:
            jump_to: ジャンプ先
            text_jp: 選択肢テキスト（日本語）
            condition: if列の条件式
            text_en: 選択肢テキスト（英語）
            text_id: テキストID
        """
        key = self._resolve_key(jump_to)
        entry = {
            'if': condition,
            'action': 'choice',
            'jump': key,
            'text_JP': text_jp,
            'text_EN': text_en or text_jp,
        }
        if text_id:
            entry['id'] = text_id
        self.entries.append(entry)
        return self

    # ============================================================================
    # CWL 拡張機能: 条件ヘルパー
    # ============================================================================

    @staticmethod
    def cond_has_flag(flag_name: str) -> str:
        """hasFlag条件を生成"""
        return f"hasFlag,{flag_name}"

    @staticmethod
    def cond_no_flag(flag_name: str) -> str:
        """!hasFlag条件を生成"""
        return f"!hasFlag,{flag_name}"

    @staticmethod
    def cond_has_item(item_id: str) -> str:
        """hasItem条件を生成"""
        return f"hasItem,{item_id}"

    @staticmethod
    def cond_quest_completed(quest_id: str) -> str:
        """isCompleted条件を生成"""
        return f"isCompleted,{quest_id}"

    @staticmethod
    def cond_flag_equals(flag_name: str, value: int) -> str:
        """フラグ値が一致する条件を生成"""
        return f"=,{flag_name},{value}"

    @staticmethod
    def cond_flag_greater(flag_name: str, value: int) -> str:
        """フラグ値が大きい条件を生成"""
        return f">,{flag_name},{value}"

    @staticmethod
    def cond_flag_less(flag_name: str, value: int) -> str:
        """フラグ値が小さい条件を生成"""
        return f"<,{flag_name},{value}"

    # ============================================================================
    # CWL 拡張機能: サウンド
    # ============================================================================

    def play_sound_with_probability(self, sound_id: str, probability: float = 1.0) -> 'DramaBuilder':
        """
        確率付き効果音を再生

        Args:
            sound_id: サウンドID
            probability: 再生確率 (0.0-1.0)
        """
        if probability >= 1.0:
            self.entries.append({'action': 'sound', 'param': sound_id})
        else:
            self.entries.append({'action': 'sound', 'param': f'{sound_id},{probability}'})
        return self

    @staticmethod
    def sound_tag(sound_id: str, probability: float = 1.0) -> str:
        """
        テキスト内に埋め込むサウンドタグを生成

        Args:
            sound_id: サウンドID
            probability: 再生確率 (0.0-1.0)

        Returns:
            埋め込み用タグ文字列 (例: "<sound=gandalf,0.8>")
        """
        if probability >= 1.0:
            return f"<sound={sound_id}>"
        return f"<sound={sound_id},{probability}>"

    # ============================================================================
    # CWL 拡張機能: 視覚効果
    # ============================================================================

    def fade_in(self, color: str = "black") -> 'DramaBuilder':
        """フェードイン（画面が明るくなる）"""
        self.entries.append({'action': 'fadeIn', 'param': color})
        return self

    def fade_out(self, color: str = "black") -> 'DramaBuilder':
        """フェードアウト（画面が暗くなる）"""
        self.entries.append({'action': 'fadeOut', 'param': color})
        return self

    def fade_in_out(self, color: str = "black") -> 'DramaBuilder':
        """フェードイン・アウト"""
        self.entries.append({'action': 'fadeInOut', 'param': color})
        return self

    def alpha_in(self) -> 'DramaBuilder':
        """アルファフェードイン"""
        self.entries.append({'action': 'alphaIn'})
        return self

    def alpha_out(self) -> 'DramaBuilder':
        """アルファフェードアウト"""
        self.entries.append({'action': 'alphaOut'})
        return self

    def shake(self) -> 'DramaBuilder':
        """画面を揺らす"""
        self.entries.append({'action': 'shake'})
        return self

    def hide_ui(self) -> 'DramaBuilder':
        """UIを非表示"""
        self.entries.append({'action': 'hideUI'})
        return self

    def hide_dialog(self) -> 'DramaBuilder':
        """ダイアログを非表示"""
        self.entries.append({'action': 'hideDialog'})
        return self

    # ============================================================================
    # CWL 拡張機能: アイテム・リソース
    # ============================================================================

    def add_key_item(self, key_item_id: str) -> 'DramaBuilder':
        """キーアイテムを追加"""
        self.entries.append({'action': 'addKeyItem', 'param': key_item_id})
        return self

    def drop_item(self, item_id: str, count: int = 1) -> 'DramaBuilder':
        """アイテムをドロップ"""
        param = f"{item_id},{count}" if count > 1 else item_id
        self.entries.append({'action': 'drop', 'param': param})
        return self

    def add_resource(self, resource_type: str, amount: int) -> 'DramaBuilder':
        """リソースを追加 (ホームリソースなど)"""
        self.entries.append({'action': 'addResource', 'param': f'{resource_type},{amount}'})
        return self

    def destroy_item(self, item_id: str) -> 'DramaBuilder':
        """アイテムを破壊"""
        self.entries.append({'action': 'destroyItem', 'param': item_id})
        return self

    # ============================================================================
    # CWL 拡張機能: カメラ・フォーカス
    # ============================================================================

    def focus_chara(self, chara_id: str, wait_before: float = 0.3, wait_after: float = 0.3) -> 'DramaBuilder':
        """
        キャラクターにフォーカス（前後にウェイト付き）

        Args:
            chara_id: キャラクターID
            wait_before: フォーカス前の待機秒数
            wait_after: フォーカス後の待機秒数
        """
        if wait_before > 0:
            self.entries.append({'action': 'wait', 'param': str(wait_before)})
        self.entries.append({'action': 'focusChara', 'param': chara_id})
        if wait_after > 0:
            self.entries.append({'action': 'wait', 'param': str(wait_after)})
        return self

    def focus_pc(self, wait_before: float = 0.3, wait_after: float = 0.3) -> 'DramaBuilder':
        """
        プレイヤーキャラにフォーカス（前後にウェイト付き）

        Args:
            wait_before: フォーカス前の待機秒数
            wait_after: フォーカス後の待機秒数
        """
        if wait_before > 0:
            self.entries.append({'action': 'wait', 'param': str(wait_before)})
        self.entries.append({'action': 'focusPC'})
        if wait_after > 0:
            self.entries.append({'action': 'wait', 'param': str(wait_after)})
        return self

    def unfocus(self, wait_before: float = 0.2, wait_after: float = 0.2) -> 'DramaBuilder':
        """
        フォーカス解除（前後にウェイト付き）

        Args:
            wait_before: 解除前の待機秒数
            wait_after: 解除後の待機秒数
        """
        if wait_before > 0:
            self.entries.append({'action': 'wait', 'param': str(wait_before)})
        self.entries.append({'action': 'unfocus'})
        if wait_after > 0:
            self.entries.append({'action': 'wait', 'param': str(wait_after)})
        return self

    # ============================================================================
    # CWL 拡張機能: システム
    # ============================================================================

    def save_game(self) -> 'DramaBuilder':
        """ゲームをセーブ"""
        self.entries.append({'action': 'save'})
        return self

    def set_hour(self, hour: int) -> 'DramaBuilder':
        """時刻を設定"""
        self.entries.append({'action': 'setHour', 'param': str(hour)})
        return self

    def build(self) -> List[Dict[str, Any]]:
        """エントリーリストを返す（検証なし）"""
        return self.entries

    def save(self, filepath: str, sheet_name: str = "main") -> None:
        """
        CWL準拠のExcelファイルとして保存。

        Args:
            filepath: 出力ファイルパス
            sheet_name: シート名（キャラクターIDなどを推奨）
        """
        os.makedirs(os.path.dirname(filepath), exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Row 1: Headers
        for col, header in enumerate(HEADERS, 1):
            ws.cell(row=1, column=col, value=header)

        # Row 2-5: 空行（CWL形式に準拠）

        # Row 6+: Data
        row = 6
        for entry in self.entries:
            for col, header in enumerate(HEADERS, 1):
                value = entry.get(header)
                if value is not None:
                    ws.cell(row=row, column=col, value=value)
            row += 1

        wb.save(filepath)
        print(f"Created: {filepath} (sheet: {sheet_name})")
