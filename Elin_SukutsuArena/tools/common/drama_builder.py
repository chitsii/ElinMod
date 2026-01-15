"""
DramaBuilder - CWL ドラマファイル生成用ビルダークラス

CWL (Custom Whatever Loader) のドラマファイル形式に準拠した
Excel ファイルを生成するためのビルダーパターン実装。

フォーマット仕様は docs/cwl_drama_format.md を参照。
"""

import openpyxl
import os
from typing import Optional, List, Union, Dict, Any, Set

# フラグ定義をインポート（検証用）
try:
    from flag_definitions import (
        get_all_flags, get_all_enums,
        EnumFlag, IntFlag, BoolFlag, StringFlag,
        PREFIX as FLAG_PREFIX,
        Actors, FlagValues, Keys
    )
    FLAG_VALIDATION_ENABLED = True
except ImportError:
    FLAG_VALIDATION_ENABLED = False
    Actors = None
    FlagValues = None
    Keys = None
    print("[DramaBuilder] Warning: flag_definitions not found, flag validation disabled")


def _validate_flag(flag_key: str, value: Any) -> List[str]:
    """
    フラグキーと値を検証。
    エラーがあればメッセージのリストを返す。

    Args:
        flag_key: フラグキー（例: "chitsii.arena.player.motivation"）
        value: 設定する値

    Returns:
        エラーメッセージのリスト（空なら問題なし）
    """
    if not FLAG_VALIDATION_ENABLED:
        return []

    errors = []

    # プレフィクスを持つフラグのみ検証
    if not flag_key.startswith(FLAG_PREFIX + "."):
        # 旧来のフラグ（sukutsu_gladiator等）はスキップ
        return []

    # 登録されたフラグか確認
    all_flags = {f.full_key: f for f in get_all_flags()}

    if flag_key not in all_flags:
        errors.append(f"Unknown flag key: '{flag_key}'")
        return errors

    flag_def = all_flags[flag_key]

    # 型チェック
    if isinstance(flag_def, EnumFlag):
        if flag_def.enum_type:
            valid_ordinals = list(range(len(flag_def.enum_type)))
            if isinstance(value, int) and value not in valid_ordinals and value != -1:
                errors.append(
                    f"Invalid ordinal {value} for enum flag '{flag_key}'. "
                    f"Valid: {valid_ordinals} or -1 (null)"
                )
    elif isinstance(flag_def, IntFlag):
        if not isinstance(value, int):
            errors.append(f"Flag '{flag_key}' expects int, got {type(value).__name__}")
        elif hasattr(flag_def, 'min_value') and hasattr(flag_def, 'max_value'):
            if value < flag_def.min_value or value > flag_def.max_value:
                errors.append(
                    f"Value {value} out of range for '{flag_key}'. "
                    f"Expected: [{flag_def.min_value}, {flag_def.max_value}]"
                )
    elif isinstance(flag_def, BoolFlag):
        if value not in (0, 1, True, False):
            errors.append(f"Flag '{flag_key}' expects bool (0/1), got {value}")

    return errors

# CWL準拠ヘッダー (if + if2 で複合条件、version/text列あり)
HEADERS = ['step', 'jump', 'if', 'if2', 'action', 'param', 'actor', 'version', 'id', 'text_JP', 'text_EN', 'text']


class DramaLabel:
    """ステップラベルを表すクラス"""
    def __init__(self, key: str):
        self.key = key
    def __str__(self):
        return self.key


class DramaActor:
    """アクターを表すクラス"""
    def __init__(self, key: str):
        self.key = key
    def __str__(self):
        return self.key


class ChoiceReaction:
    """
    選択肢とその反応を一体化して定義するクラス。
    choice_block()と組み合わせて使用する。

    使用例:
        builder.choice_block([
            ChoiceReaction("闘士になりたい", "c1")
                .say("join1", "おまえが？ハーッハッハ...", actor=vargus)
                .set_flag("sukutsu_gladiator", 1)
                .jump(registered),

            ChoiceReaction("いや、やめておく", "c2")
                .say("reject1", "話は終わりだ。", actor=vargus)
                .jump(end),
        ], cancel=end)
    """

    def __init__(self, text_jp: str, text_id: str = "", text_en: str = "", condition: str = ""):
        """
        Args:
            text_jp: 選択肢テキスト（日本語）
            text_id: テキストID（省略可）
            text_en: 選択肢テキスト（英語、省略時は日本語と同じ）
            condition: 表示条件（if列、省略可）
        """
        self.text_jp = text_jp
        self.text_en = text_en or text_jp
        self.text_id = text_id
        self.condition = condition
        self.actions: List[Dict[str, Any]] = []
        self._label: Optional[str] = None  # choice_blockで自動設定

    def _add_action(self, action: Dict[str, Any]) -> 'ChoiceReaction':
        """アクションを追加（内部用）"""
        self.actions.append(action)
        return self

    # === 会話系 ===

    def say(self, text_id: str, text_jp: str, text_en: str = "",
            actor: Union[str, 'DramaActor'] = None) -> 'ChoiceReaction':
        """テキスト行を追加"""
        actor_key = actor.key if isinstance(actor, DramaActor) else actor
        entry = {
            'id': text_id,
            'text_JP': text_jp,
            'text_EN': text_en or text_jp,
        }
        if actor_key:
            entry['actor'] = actor_key
        return self._add_action(entry)

    # === フロー制御 ===

    def jump(self, jump_to: Union[str, 'DramaLabel']) -> 'ChoiceReaction':
        """指定ステップにジャンプ"""
        key = jump_to.key if isinstance(jump_to, DramaLabel) else jump_to
        return self._add_action({'jump': key})

    def end(self) -> 'ChoiceReaction':
        """ドラマを終了"""
        return self._add_action({'action': 'end'})

    # === フラグ操作 ===

    def set_flag(self, flag: str, value: int = 1) -> 'ChoiceReaction':
        """フラグを設定"""
        return self._add_action({
            'action': 'setFlag',
            'param': f'{flag},{value}',
        })

    def mod_flag(self, flag: str, operator: str, value: int,
                 actor: Union[str, 'DramaActor'] = None) -> 'ChoiceReaction':
        """フラグを変更"""
        actor_key = actor.key if isinstance(actor, DramaActor) else (actor or 'pc')
        return self._add_action({
            'action': 'invoke*',
            'param': f'mod_flag({flag}, {operator}{value})',
            'actor': actor_key,
        })

    # === 演出 ===

    def shake(self) -> 'ChoiceReaction':
        """画面を揺らす"""
        return self._add_action({'action': 'shake'})

    def wait(self, seconds: float) -> 'ChoiceReaction':
        """待機"""
        return self._add_action({'action': 'wait', 'param': str(seconds)})

    def play_bgm(self, bgm_id: str) -> 'ChoiceReaction':
        """BGMを再生"""
        code = f'''
            Debug.Log("[SukutsuArena] Attempting to play BGM: {bgm_id}");
            var data = SoundManager.current.GetData("{bgm_id}");
            if (data != null) {{
                if (data is BGMData bgm) {{
                    LayerDrama.haltPlaylist = true;
                    LayerDrama.maxBGMVolume = true;
                    SoundManager.current.PlayBGM(bgm);
                }} else {{
                    SoundManager.current.Play(data);
                }}
            }}
        '''.replace('\n', ' ').strip()
        return self._add_action({'action': 'eval', 'param': code})

    def play_sound(self, sound_id: str) -> 'ChoiceReaction':
        """効果音を再生"""
        return self._add_action({'action': 'sound', 'param': sound_id})

    # === クエスト ===

    def complete_quest(self, quest_id: str, actor: Union[str, 'DramaActor'] = None) -> 'ChoiceReaction':
        """クエストを完了"""
        actor_key = actor.key if isinstance(actor, DramaActor) else (actor or 'pc')
        return self._add_action({
            'action': 'modInvoke',
            'param': f'complete_quest({quest_id})',
            'actor': actor_key,
        })

    # === 汎用 ===

    def action(self, action_name: str, param: str = None,
               jump: Union[str, 'DramaLabel'] = None,
               actor: Union[str, 'DramaActor'] = None) -> 'ChoiceReaction':
        """汎用アクションを追加"""
        entry = {'action': action_name}
        if param:
            entry['param'] = param
        if jump:
            entry['jump'] = jump.key if isinstance(jump, DramaLabel) else jump
        if actor:
            entry['actor'] = actor.key if isinstance(actor, DramaActor) else actor
        return self._add_action(entry)

    def eval(self, code: str, actor: Union[str, 'DramaActor'] = None) -> 'ChoiceReaction':
        """C#コードを実行"""
        entry = {'action': 'eval', 'param': code}
        if actor:
            entry['actor'] = actor.key if isinstance(actor, DramaActor) else actor
        return self._add_action(entry)


class DramaBuilder:
    """
    CWLドラマファイルを構築するビルダークラス。

    使用例:
        drama = DramaBuilder()
        pc = drama.register_actor("pc", "プレイヤー", "Player")
        npc = drama.register_actor("master", "マスター", "Master")

        lbl_main = drama.label("main")
        drama.step(lbl_main)
        drama.say("greet", "こんにちは", actor=npc)
        drama.finish()

        drama.save("drama.xlsx", sheet_name="master")
    """

    def __init__(self) -> None:
        self.entries: List[Dict[str, Any]] = []
        self.actors: Dict[str, Dict[str, str]] = {}
        self.registered_steps: Set[str] = set()
        self._current_step: Optional[str] = None

    def _resolve_key(self, obj: Union[str, DramaLabel, DramaActor]) -> str:
        """オブジェクトからキーを取得"""
        if isinstance(obj, (DramaLabel, DramaActor)):
            return obj.key
        return obj

    def register_actor(self, actor_id: str, name_jp: str, name_en: str = "") -> DramaActor:
        """アクターを登録"""
        self.actors[actor_id] = {'jp': name_jp, 'en': name_en or name_jp}
        return DramaActor(actor_id)

    def label(self, key: str) -> DramaLabel:
        """ラベルを作成"""
        return DramaLabel(key)

    def step(self, step_label: Union[str, DramaLabel]) -> 'DramaBuilder':
        """
        新しいステップを開始。
        CWL形式では、step行は step列のみを含み、内容は次の行から。
        """
        key = self._resolve_key(step_label)
        self.registered_steps.add(key)
        self._current_step = key
        # step行を追加（step列のみ）
        self.entries.append({'step': key})
        return self

    def say(self, text_id: str, text_jp: str, text_en: str = "",
            actor: Union[str, DramaActor] = None) -> 'DramaBuilder':
        """テキスト行を追加"""
        actor_key = self._resolve_key(actor) if actor else None
        entry = {
            'id': text_id,
            'text_JP': text_jp,
            'text_EN': text_en or text_jp,
        }
        if actor_key:
            entry['actor'] = actor_key
        self.entries.append(entry)
        return self

    def choice(self, jump_to: Union[str, DramaLabel], text_jp: str,
               text_en: str = "", text_id: str = "", condition: str = "") -> 'DramaBuilder':
        """選択肢を追加"""
        key = self._resolve_key(jump_to)
        entry = {
            'action': 'choice',
            'jump': key,
            'text_JP': text_jp,
            'text_EN': text_en or text_jp,
        }
        if text_id:
            entry['id'] = text_id
        if condition:
            entry['if'] = condition
        self.entries.append(entry)
        return self

    def jump(self, jump_to: Union[str, DramaLabel]) -> 'DramaBuilder':
        """指定ステップにジャンプ"""
        key = self._resolve_key(jump_to)
        self.entries.append({'jump': key})
        return self

    def branch_if(self, flag: str, operator: str, value: int,
                  jump_to: Union[str, DramaLabel],
                  actor: Union[str, DramaActor] = None) -> 'DramaBuilder':
        """
        条件分岐。invoke* + if_flag を使用。

        注意: 複数のbranch_ifを連続して使用する場合、最後のフォールバックjumpが
        先に実行される可能性があります。代わりに switch_on_flag() を使用してください。
        """
        key = self._resolve_key(jump_to)
        actor_key = self._resolve_key(actor) if actor else 'pc'
        entry = {
            'action': 'invoke*',
            'param': f'if_flag({flag}, {operator}{value})',
            'jump': key,
            'actor': actor_key,
        }
        self.entries.append(entry)
        return self

    def branch_quest_done(self, quest_id: str, jump_to: Union[str, DramaLabel],
                          actor: Union[str, DramaActor] = None) -> 'DramaBuilder':
        """
        クエスト完了済みなら指定ラベルにジャンプ。

        Args:
            quest_id: クエストID（例: "07_upper_existence"）
            jump_to: 完了済みの場合のジャンプ先
            actor: アクター（デフォルト: pc）

        Note:
            クエスト完了フラグは "sukutsu_quest_done_{quest_id}" の形式で保存されている。
        """
        flag_key = f"sukutsu_quest_done_{quest_id}"
        return self.branch_if(flag_key, "==", 1, jump_to, actor)

    def switch_on_flag(self, flag: str, cases: Dict[int, Union[str, DramaLabel]],
                       fallback: Union[str, DramaLabel] = None,
                       actor: Union[str, DramaActor] = None) -> 'DramaBuilder':
        """
        フラグ値に基づく複数条件分岐（安全なswitch-case風API）

        modInvokeのif_flagコマンドを使用して条件分岐を実現。
        各ケースは独立して評価され、条件が一致した場合のみジャンプします。

        Args:
            flag: チェックするフラグキー
            cases: {値: ジャンプ先} の辞書
            fallback: どの条件にも一致しない場合のジャンプ先（省略可）
            actor: アクター（デフォルト: pc）

        Example:
            builder.switch_on_flag("sukutsu_quest_target_name", {
                11: start_rank_g,
                12: start_rank_f,
                16: start_rank_b,
            }, fallback=rank_up_not_ready)

        Note:
            フォールバックはフラグ値が0の場合にのみ実行されます。
            check_quest_available等でフラグが設定されなかった場合に
            0のままになるため、これが正しい動作となります。
        """
        actor_key = self._resolve_key(actor) if actor else 'pc'

        # 各ケースをmodInvokeのif_flagとして生成
        # ジャンプ先はjumpカラムで指定（CWLのjumpFuncパターンに対応）
        for value, jump_to in cases.items():
            key = self._resolve_key(jump_to)
            entry = {
                'action': 'modInvoke',
                'param': f'if_flag({flag}, =={value})',
                'jump': key,
                'actor': actor_key,
            }
            self.entries.append(entry)

        # フォールバックも条件付きで生成（値が0の場合のみ）
        if fallback is not None:
            fallback_key = self._resolve_key(fallback)
            entry = {
                'action': 'modInvoke',
                'param': f'if_flag({flag}, ==0)',
                'jump': fallback_key,
                'actor': actor_key,
            }
            self.entries.append(entry)

        return self

    def switch_flag(self, flag: str, cases: List[Union[str, DramaLabel]],
                    fallback: Union[str, DramaLabel] = None,
                    actor: Union[str, DramaActor] = None) -> 'DramaBuilder':
        """
        フラグ値に基づいて直接ジャンプ（switch_flagコマンド使用）

        switch_on_flagと異なり、drama.sequence.Play()を直接呼び出す。
        jumpFuncに依存しないため、より確実に動作する。

        Args:
            flag: チェックするフラグキー
            cases: [flag=0のジャンプ先, flag=1のジャンプ先, ...]
            fallback: どれにも該当しない場合のジャンプ先（省略可）
            actor: アクター（デフォルト: pc）

        Example:
            builder.switch_flag("chitsii.arena.player.rank", [
                rank_labels["start_rank_g"],  # rank=0 (unranked)
                rank_labels["start_rank_f"],  # rank=1 (G)
                rank_labels["start_rank_e"],  # rank=2 (F)
            ], fallback=rank_up_not_ready)
        """
        actor_key = self._resolve_key(actor) if actor else 'pc'
        jump_targets = [self._resolve_key(c) for c in cases]
        if fallback:
            jump_targets.append(self._resolve_key(fallback))

        self.entries.append({
            'action': 'modInvoke',
            'param': f'switch_flag({flag}, {", ".join(jump_targets)})',
            'actor': actor_key,
        })
        return self

    def choice_block(self, choices: List['ChoiceReaction'],
                     cancel: Union[str, DramaLabel] = None,
                     label_prefix: str = None) -> 'DramaBuilder':
        """
        選択肢と反応を一体化して定義する。
        ChoiceReactionオブジェクトのリストを受け取り、自動的にステップを生成。

        Args:
            choices: ChoiceReactionオブジェクトのリスト
            cancel: キャンセル時のジャンプ先（省略可）
            label_prefix: 自動生成ラベルのプレフィックス（省略時は現在のステップ名を使用）

        Example:
            builder.step("opening") \\
                .say("greet1", "何の用だ、ひよっこ...", actor=vargus) \\
                .choice_block([
                    ChoiceReaction("闘士になりたい", "c1")
                        .say("join1", "おまえが？ハーッハッハ...", actor=vargus)
                        .set_flag("sukutsu_gladiator", 1)
                        .jump("registered"),

                    ChoiceReaction("いや、やめておく", "c2")
                        .say("reject1", "話は終わりだ。", actor=vargus)
                        .jump("end"),
                ], cancel="end")

        Note:
            各ChoiceReactionに対して自動的にラベルが生成され、
            反応のステップがchoice_block呼び出し後に追加されます。
            ラベル形式: {prefix}_react_{index} (例: opening_react_0, opening_react_1)
        """
        # プレフィックス決定
        prefix = label_prefix or self._current_step or "choice"

        # 各選択肢に対してラベルを生成し、choice行を追加
        for i, cr in enumerate(choices):
            cr._label = f"{prefix}_react_{i}"

            entry = {
                'action': 'choice',
                'jump': cr._label,
                'text_JP': cr.text_jp,
                'text_EN': cr.text_en,
            }
            if cr.text_id:
                entry['id'] = cr.text_id
            if cr.condition:
                entry['if'] = cr.condition
            self.entries.append(entry)

        # キャンセル時のジャンプ先
        if cancel is not None:
            cancel_key = self._resolve_key(cancel)
            self.entries.append({'action': 'cancel', 'jump': cancel_key})

        # 各反応のステップを追加
        for cr in choices:
            # step行
            self.entries.append({'step': cr._label})
            self.registered_steps.add(cr._label)

            # アクション行
            for action in cr.actions:
                self.entries.append(action)

        return self

    def check_quests(self, checks: list, actor: Union[str, DramaActor] = None) -> 'DramaBuilder':
        """
        複数のクエスト利用可能チェックを一括で実行。

        Args:
            checks: (quest_id, jump_label) タプルのリスト
            actor: アクター（デフォルト: pc）

        Example:
            builder.check_quests([
                (QuestIds.RANK_UP_G, start_rank_g),
                (QuestIds.RANK_UP_F, start_rank_f),
                (QuestIds.RANK_UP_E, start_rank_e),
            ])
        """
        for quest_id, jump_target in checks:
            self.check_quest_available(quest_id, jump_target, actor)
        return self

    def set_flag(self, flag: str, value: int = 1) -> 'DramaBuilder':
        """フラグを設定（検証付き）"""
        # フラグ検証
        errors = _validate_flag(flag, value)
        if errors:
            for err in errors:
                self._add_validation_error(f"set_flag: {err}")

        self.entries.append({
            'action': 'setFlag',
            'param': f'{flag},{value}',
        })
        return self

    def mod_flag(self, flag: str, operator: str, value: int,
                 actor: Union[str, DramaActor] = None) -> 'DramaBuilder':
        """フラグを変更 (invoke* mod_flag) - 検証付き"""
        # mod_flag の場合、絶対値で検証（加減算なので範囲チェックは緩め）
        if operator == "=" :
            errors = _validate_flag(flag, value)
            if errors:
                for err in errors:
                    self._add_validation_error(f"mod_flag: {err}")

        actor_key = self._resolve_key(actor) if actor else 'pc'
        self.entries.append({
            'action': 'invoke*',
            'param': f'mod_flag({flag}, {operator}{value})',
            'actor': actor_key,
        })
        return self

    def eval(self, code: str, actor: Union[str, DramaActor] = None) -> 'DramaBuilder':
        """C#コードを実行"""
        entry = {'action': 'eval', 'param': code}
        if actor:
            entry['actor'] = self._resolve_key(actor)
        self.entries.append(entry)
        return self

    def _add_validation_error(self, message: str) -> None:
        """検証エラーを追加"""
        if not hasattr(self, '_validation_errors'):
            self._validation_errors = []
        self._validation_errors.append(message)
        print(f"[DramaBuilder] ERROR: {message}")

    def get_validation_errors(self) -> List[str]:
        """検証エラーを取得"""
        return getattr(self, '_validation_errors', [])

    def action(self, action_name: str, param: str = None,
               jump: Union[str, DramaLabel] = None,
               actor: Union[str, DramaActor] = None) -> 'DramaBuilder':
        """汎用アクションを追加"""
        entry = {'action': action_name}
        if param:
            entry['param'] = param
        if jump:
            entry['jump'] = self._resolve_key(jump)
        if actor:
            entry['actor'] = self._resolve_key(actor)
        self.entries.append(entry)
        return self

    def on_cancel(self, jump_to: Union[str, DramaLabel]) -> 'DramaBuilder':
        """キャンセル時の動作を設定"""
        key = self._resolve_key(jump_to)
        self.entries.append({'action': 'cancel', 'jump': key})
        return self

    def finish(self) -> 'DramaBuilder':
        """ドラマを終了"""
        self.entries.append({'action': 'end'})
        return self

    def play_sound(self, sound_id: str) -> 'DramaBuilder':
        """効果音を再生"""
        self.entries.append({'action': 'sound', 'param': sound_id})
        return self

    def play_bgm(self, bgm_id: str) -> 'DramaBuilder':
        """
        BGMを再生

        Args:
            bgm_id: BGM ID (例: "BGM/sukutsu_arena_opening")

        Note:
            CWLカスタムBGMは SoundManager.current.GetData(id) で取得して再生
        """
        # evalアクションでC#コードを実行（デバッグログ付き）
        # CWLはSoundManager.current.dictDataにサウンドを登録し、
        # GetData(id)で取得可能。BGMはBGMDataとして作成される。
        code = f'''
            Debug.Log("[SukutsuArena] Attempting to play BGM: {bgm_id}");
            var data = SoundManager.current.GetData("{bgm_id}");
            if (data != null) {{
                Debug.Log("[SukutsuArena] Found BGM data, type: " + data.GetType().Name);
                if (data is BGMData bgm) {{
                    Debug.Log("[SukutsuArena] Playing as BGM");
                    LayerDrama.haltPlaylist = true;
                    LayerDrama.maxBGMVolume = true;
                    SoundManager.current.PlayBGM(bgm);
                }} else {{
                    Debug.Log("[SukutsuArena] Playing as Sound");
                    SoundManager.current.Play(data);
                }}
            }} else {{
                Debug.LogWarning("[SukutsuArena] BGM not found: {bgm_id}");
            }}
        '''.replace('\n', ' ').strip()
        self.entries.append({
            'action': 'eval',
            'param': code
        })
        return self

    def play_bgm_vanilla(self, bgm_id: int) -> 'DramaBuilder':
        """
        バニラBGMを再生（数値ID）

        Args:
            bgm_id: バニラBGM ID (数値)
        """
        code = f'''
            Debug.Log("[SukutsuArena] Playing vanilla BGM ID: {bgm_id}");
            if (EMono.core.refs.dictBGM.TryGetValue({bgm_id}, out var bgm)) {{
                LayerDrama.haltPlaylist = true;
                LayerDrama.maxBGMVolume = true;
                EMono.Sound.PlayBGM(bgm);
            }} else {{
                Debug.LogWarning("[SukutsuArena] Vanilla BGM not found: {bgm_id}");
            }}
        '''.replace('\n', ' ').strip()
        self.entries.append({
            'action': 'eval',
            'param': code
        })
        return self

    def wait(self, seconds: float) -> 'DramaBuilder':
        """待機"""
        self.entries.append({'action': 'wait', 'param': str(seconds)})
        return self

    def effect(self, effect_id: str) -> 'DramaBuilder':
        """エフェクトを再生"""
        self.entries.append({'action': effect_id})
        return self

    # ============================================================================
    # CWL 拡張機能: 組み込みジャンプ
    # ============================================================================

    def inject_unique(self) -> 'DramaBuilder':
        """
        inject/Unique アクションを実行
        標準的な会話UIを挿入（パーティ加入、商品購入など）
        """
        self.entries.append({'action': 'inject', 'param': 'Unique'})
        return self

    def jump_to_trade(self) -> 'DramaBuilder':
        """商人との取引画面にジャンプ (actor需要はtgが対象)"""
        self.entries.append({'jump': '_trade'})
        return self

    def jump_to_buy(self) -> 'DramaBuilder':
        """購入画面にジャンプ"""
        self.entries.append({'jump': '_buy'})
        return self

    def jump_to_join_party(self) -> 'DramaBuilder':
        """パーティ加入にジャンプ"""
        self.entries.append({'jump': '_joinParty'})
        return self

    def jump_to_leave_party(self) -> 'DramaBuilder':
        """パーティ離脱にジャンプ"""
        self.entries.append({'jump': '_leaveParty'})
        return self

    def jump_to_train(self) -> 'DramaBuilder':
        """訓練画面にジャンプ"""
        self.entries.append({'jump': '_train'})
        return self

    def jump_to_heal(self) -> 'DramaBuilder':
        """回復サービスにジャンプ"""
        self.entries.append({'jump': '_heal'})
        return self

    def jump_to_invest_shop(self) -> 'DramaBuilder':
        """店への投資にジャンプ"""
        self.entries.append({'jump': '_investShop'})
        return self

    def jump_to_sell_fame(self) -> 'DramaBuilder':
        """名声売却にジャンプ"""
        self.entries.append({'jump': '_sellFame'})
        return self

    def jump_to_copy_item(self) -> 'DramaBuilder':
        """アイテム複製にジャンプ"""
        self.entries.append({'jump': '_copyItem'})
        return self

    def jump_to_give(self) -> 'DramaBuilder':
        """アイテム渡すにジャンプ"""
        self.entries.append({'jump': '_give'})
        return self

    def jump_to_whore(self) -> 'DramaBuilder':
        """売春サービスにジャンプ (18+)"""
        self.entries.append({'jump': '_whore'})
        return self

    def jump_to_tail(self) -> 'DramaBuilder':
        """尾行にジャンプ"""
        self.entries.append({'jump': '_tail'})
        return self

    def jump_to_suck(self) -> 'DramaBuilder':
        """吸血にジャンプ (サキュバス等)"""
        self.entries.append({'jump': '_suck'})
        return self

    def jump_to_bout(self) -> 'DramaBuilder':
        """決闘にジャンプ"""
        self.entries.append({'jump': '_bout'})
        return self

    def jump_to_rumor(self) -> 'DramaBuilder':
        """噂話にジャンプ"""
        self.entries.append({'jump': '_rumor'})
        return self

    def jump_to_news(self) -> 'DramaBuilder':
        """ニュースにジャンプ"""
        self.entries.append({'jump': '_news'})
        return self

    def jump_builtin(self, builtin_name: str) -> 'DramaBuilder':
        """
        汎用組み込みジャンプ

        Args:
            builtin_name: 組み込みステップ名 (例: "_rumor", "_bout", "_news")
        """
        if not builtin_name.startswith('_'):
            builtin_name = '_' + builtin_name
        self.entries.append({'jump': builtin_name})
        return self

    # ============================================================================
    # CWL 拡張機能: ランタイムスクリプト
    # ============================================================================

    def cs_eval(self, code: str, actor: Union[str, DramaActor] = None) -> 'DramaBuilder':
        """
        C# コードをランタイムで実行 (eval アクション)

        Args:
            code: 実行するC#コード
            actor: 対象アクター (デフォルト: tg)

        Note:
            - Script["var"] で変数を複数 eval 間で共有可能
            - 複雑なロジックは DLL に実装することを推奨
        """
        actor_key = self._resolve_key(actor) if actor else None
        entry = {
            'action': 'eval',
            'param': code,
        }
        if actor_key:
            entry['actor'] = actor_key
        self.entries.append(entry)
        return self

    def cs_script_var_set(self, var_name: str, value_expr: str) -> 'DramaBuilder':
        """
        スクリプト共有変数を設定

        Args:
            var_name: 変数名
            value_expr: C#式として評価される値
        """
        return self.cs_eval(f'Script["{var_name}"] = {value_expr};')

    def cs_script_var_get(self, var_name: str, cast_type: str = "int") -> str:
        """
        スクリプト共有変数を取得するC#式を生成

        Args:
            var_name: 変数名
            cast_type: キャスト型 (例: "int", "string", "bool")

        Returns:
            C#式文字列 (例: "(int)Script[\"counter\"]")
        """
        return f'({cast_type})Script["{var_name}"]'

    # ============================================================================
    # CWL 拡張機能: 条件付きエントリ
    # ============================================================================

    def say_if(self, text_id: str, text_jp: str, condition: str,
               text_en: str = "", actor: Union[str, DramaActor] = None) -> 'DramaBuilder':
        """
        条件付きテキスト行を追加

        Args:
            text_id: テキストID
            text_jp: 日本語テキスト
            condition: if列の条件式 (例: "hasFlag,sukutsu_gladiator" or "=,flag,1")
            text_en: 英語テキスト
            actor: 発言者
        """
        actor_key = self._resolve_key(actor) if actor else None
        entry = {
            'if': condition,
            'id': text_id,
            'text_JP': text_jp,
            'text_EN': text_en or text_jp,
        }
        if actor_key:
            entry['actor'] = actor_key
        self.entries.append(entry)
        return self

    def choice_if(self, jump_to: Union[str, DramaLabel], text_jp: str,
                  condition: str, text_en: str = "", text_id: str = "") -> 'DramaBuilder':
        """
        条件付き選択肢を追加

        Args:
            jump_to: ジャンプ先
            text_jp: 選択肢テキスト（日本語）
            condition: if列の条件式
            text_en: 選択肢テキスト（英語）
            text_id: テキストID
        """
        key = self._resolve_key(jump_to)
        entry = {
            'if': condition,
            'action': 'choice',
            'jump': key,
            'text_JP': text_jp,
            'text_EN': text_en or text_jp,
        }
        if text_id:
            entry['id'] = text_id
        self.entries.append(entry)
        return self

    def choice_if2(self, jump_to: Union[str, DramaLabel], text_jp: str,
                   condition1: str, condition2: str,
                   text_en: str = "", text_id: str = "") -> 'DramaBuilder':
        """
        複数条件付き選択肢を追加（if + if2 の両方を使用）

        CWLでは &での条件結合をサポートしていないため、
        複数条件が必要な場合はif列とif2列を別々に使用する。

        Args:
            jump_to: ジャンプ先
            text_jp: 選択肢テキスト（日本語）
            condition1: if列の条件式
            condition2: if2列の条件式
            text_en: 選択肢テキスト（英語）
            text_id: テキストID
        """
        key = self._resolve_key(jump_to)
        entry = {
            'if': condition1,
            'if2': condition2,
            'action': 'choice',
            'jump': key,
            'text_JP': text_jp,
            'text_EN': text_en or text_jp,
        }
        if text_id:
            entry['id'] = text_id
        self.entries.append(entry)
        return self

    # ============================================================================
    # CWL 拡張機能: 条件ヘルパー
    # ============================================================================

    @staticmethod
    def cond_has_flag(flag_name: str) -> str:
        """hasFlag条件を生成"""
        return f"hasFlag,{flag_name}"

    @staticmethod
    def cond_no_flag(flag_name: str) -> str:
        """!hasFlag条件を生成"""
        return f"!hasFlag,{flag_name}"

    @staticmethod
    def cond_has_item(item_id: str) -> str:
        """hasItem条件を生成"""
        return f"hasItem,{item_id}"

    @staticmethod
    def cond_quest_completed(quest_id: str) -> str:
        """isCompleted条件を生成"""
        return f"isCompleted,{quest_id}"

    @staticmethod
    def cond_flag_equals(flag_name: str, value: int) -> str:
        """フラグ値が一致する条件を生成"""
        return f"=,{flag_name},{value}"

    @staticmethod
    def cond_flag_greater(flag_name: str, value: int) -> str:
        """フラグ値が大きい条件を生成"""
        return f">,{flag_name},{value}"

    @staticmethod
    def cond_flag_less(flag_name: str, value: int) -> str:
        """フラグ値が小さい条件を生成"""
        return f"<,{flag_name},{value}"

    # ============================================================================
    # CWL 拡張機能: 視覚効果
    # ============================================================================

    def fade_in(self, duration: float = 1.0, color: str = "black") -> 'DramaBuilder':
        """
        フェードイン（画面が明るくなる）

        Args:
            duration: フェード時間（秒）
            color: フェード色 ("black" または "white")
        """
        # 色を常に第2パラメータとして明示的に指定
        param = f"{duration},{color}"
        self.entries.append({'action': 'fadeIn', 'param': param})
        return self

    def fade_out(self, duration: float = 1.0, color: str = "black") -> 'DramaBuilder':
        """
        フェードアウト（画面が暗くなる）

        Args:
            duration: フェード時間（秒）
            color: フェード色 ("black" または "white")
        """
        # 色を常に第2パラメータとして明示的に指定
        param = f"{duration},{color}"
        self.entries.append({'action': 'fadeOut', 'param': param})
        return self

    def set_background(self, bg_id: str) -> 'DramaBuilder':
        """
        背景画像を設定（CWL eval経由でTextureフォルダから読み込み）

        Args:
            bg_id: 背景画像ID（Texture/{bg_id}.png）
        """
        # CWLのSpriteCreator拡張メソッドを使用
        code = f'dm.imageBG.enabled = true; dm.imageBG.sprite = "{bg_id}".LoadSprite();'
        self.entries.append({'action': 'eval', 'param': code})
        return self

    def glitch(self) -> 'DramaBuilder':
        """グリッチエフェクトを有効化"""
        self.entries.append({'action': 'glitch'})
        return self

    def shake(self) -> 'DramaBuilder':
        """画面を揺らす"""
        self.entries.append({'action': 'shake'})
        return self

    def set_dialog_style(self, style: str = "Default") -> 'DramaBuilder':
        """
        ダイアログスタイルを変更

        Args:
            style: ダイアログスタイル
                - "Default": 標準ダイアログ
                - "Default2": 別スタイル
                - "Mono": モノローグ（キャラ名なし、グレースケール背景）
        """
        self.entries.append({'action': 'setDialog', 'param': f",,{style}"})
        return self

    # ============================================================================
    # ハイレベルAPI: ドラマ演出
    # ============================================================================

    def drama_start(
        self,
        bg_id: str = None,
        bgm_id: str = None,
        fade_duration: float = 1.0
    ) -> 'DramaBuilder':
        """
        ドラマ開始演出
        fadeOut → set_background(オプション) → fadeIn → play_bgm(オプション)

        Args:
            bg_id: 背景画像ID (Texture/フォルダからの相対パス、例: "Drama/arena_lobby")
            bgm_id: BGM ID (例: "BGM/sukutsu_arena_opening")
            fade_duration: フェード時間（秒）
        """
        self.fade_out(duration=fade_duration, color="black")
        if bg_id:
            self.set_background(bg_id)
        self.fade_in(duration=fade_duration, color="black")
        if bgm_id:
            self.play_bgm(bgm_id)
        return self

    def drama_end(self, fade_duration: float = 1.0) -> 'DramaBuilder':
        """
        ドラマ終了演出
        fadeOut → finish

        Args:
            fade_duration: フェード時間（秒）
        """
        self.fade_out(duration=fade_duration, color="black")
        self.finish()
        return self

    def scene_transition(
        self,
        bg_id: str = None,
        bgm_id: str = None,
        fade_duration: float = 0.5
    ) -> 'DramaBuilder':
        """
        シーン切り替え演出
        fadeOut → set_background(オプション) → fadeIn → play_bgm(オプション)

        Args:
            bg_id: 新しい背景画像ID
            bgm_id: 新しいBGM ID (省略時は継続)
            fade_duration: フェード時間（秒）
        """
        self.fade_out(duration=fade_duration, color="black")
        if bg_id:
            self.set_background(bg_id)
        self.fade_in(duration=fade_duration, color="black")
        if bgm_id:
            self.play_bgm(bgm_id)
        return self

    # ============================================================================
    # CWL 拡張機能: カメラ・フォーカス
    # ============================================================================

    def focus_chara(self, chara_id: str, wait_before: float = 0.3, wait_after: float = 0.5) -> 'DramaBuilder':
        """
        キャラクターにフォーカス（前後にウェイト付き）

        Args:
            chara_id: キャラクターID
            wait_before: フォーカス前の待機秒数
            wait_after: フォーカス後の待機秒数
        """
        if wait_before > 0:
            self.entries.append({'action': 'wait', 'param': str(wait_before)})
        self.entries.append({'action': 'focusChara', 'param': chara_id})
        if wait_after > 0:
            self.entries.append({'action': 'wait', 'param': str(wait_after)})
        return self

    def unfocus(self) -> 'DramaBuilder':
        """
        カメラフォーカスを解除（PCに戻す）

        ドラマ終了前に呼び出すことで、画面固定を防ぐ
        """
        self.entries.append({'action': 'unfocus'})
        return self

    # ============================================================================
    # CWL 拡張機能: システム
    # ============================================================================

    def save_game(self) -> 'DramaBuilder':
        """ゲームをセーブ"""
        self.entries.append({'action': 'save'})
        return self

    def set_hour(self, hour: int) -> 'DramaBuilder':
        """時刻を設定"""
        self.entries.append({'action': 'setHour', 'param': str(hour)})
        return self

    # ============================================================================
    # CWL 拡張機能: クエスト操作
    # ============================================================================

    def start_quest(self, quest_id: str) -> 'DramaBuilder':
        """
        新しいクエストを開始する

        Args:
            quest_id: クエストID (例: "q_collect_stone")
        """
        self.entries.append({'action': 'startQuest', 'param': quest_id})
        return self

    def complete_quest(self, quest_id: str = "") -> 'DramaBuilder':
        """
        クエストを完了する

        Args:
            quest_id: クエストID（省略時は現在のクエスト）
        """
        if quest_id:
            self.entries.append({'action': 'completeQuest', 'param': quest_id})
        else:
            self.entries.append({'action': 'completeQuest'})
        return self

    def next_phase(self, quest_id: str) -> 'DramaBuilder':
        """
        クエストの次フェーズに進む

        Args:
            quest_id: クエストID
        """
        self.entries.append({'action': 'nextPhase', 'param': quest_id})
        return self

    def change_phase(self, quest_id: str, phase: int) -> 'DramaBuilder':
        """
        クエストのフェーズを変更する

        Args:
            quest_id: クエストID
            phase: フェーズ番号
        """
        self.entries.append({'action': 'changePhase', 'param': f'{quest_id},{phase}'})
        return self

    def set_quest_client(self) -> 'DramaBuilder':
        """
        現在のクエストのクライアントを tg（会話相手）に設定する
        """
        self.entries.append({'action': 'setQuestClient'})
        return self

    def update_journal(self) -> 'DramaBuilder':
        """
        ジャーナルを更新する
        """
        self.entries.append({'action': 'updateJournal'})
        return self

    # ============================================================================
    # Arena Quest System: modInvoke Actions
    # ============================================================================

    def mod_invoke(self, method_call: str, actor: Union[str, DramaActor] = None) -> 'DramaBuilder':
        """
        C# メソッドを modInvoke アクションで呼び出す

        Args:
            method_call: メソッド呼び出し文字列（例: "check_quest_available(quest_01, label)"）
            actor: アクター（デフォルト: pc）
        """
        actor_key = self._resolve_key(actor) if actor else 'pc'
        self.entries.append({
            'action': 'modInvoke',
            'param': method_call,
            'actor': actor_key,
        })
        return self

    def check_quest_available(self, quest_id: str, jump_to: Union[str, DramaLabel],
                               actor: Union[str, DramaActor] = None) -> 'DramaBuilder':
        """
        クエストが利用可能かチェックし、利用可能ならジャンプ

        Args:
            quest_id: クエストID
            jump_to: ジャンプ先ラベル
            actor: アクター（デフォルト: pc）
        """
        jump_key = self._resolve_key(jump_to)
        return self.mod_invoke(f'check_quest_available({quest_id}, {jump_key})', actor)

    def start_quest(self, quest_id: str, actor: Union[str, DramaActor] = None) -> 'DramaBuilder':
        """
        クエストを開始

        Args:
            quest_id: クエストID
            actor: アクター（デフォルト: pc）
        """
        return self.mod_invoke(f'start_quest({quest_id})', actor)

    def complete_quest(self, quest_id: str, actor: Union[str, DramaActor] = None) -> 'DramaBuilder':
        """
        クエストを完了

        Args:
            quest_id: クエストID
            actor: アクター（デフォルト: pc）
        """
        return self.mod_invoke(f'complete_quest({quest_id})', actor)

    def if_flag_string(self, flag: str, operator: str, value: str,
                       jump_to: Union[str, DramaLabel],
                       actor: Union[str, DramaActor] = None) -> 'DramaBuilder':
        """
        フラグの文字列/enum値を比較してジャンプ（modInvoke版）

        Args:
            flag: フラグキー
            operator: 演算子（"==" または "!="）
            value: 比較する文字列値
            jump_to: 条件が真の場合のジャンプ先
            actor: アクター（デフォルト: pc）
        """
        jump_key = self._resolve_key(jump_to)
        actor_key = self._resolve_key(actor) if actor else 'pc'
        self.entries.append({
            'action': 'modInvoke',
            'param': f'if_flag({flag}, {operator}, {value}, {jump_key})',
            'actor': actor_key,
        })
        return self

    def debug_log_flags(self, actor: Union[str, DramaActor] = None) -> 'DramaBuilder':
        """
        デバッグ: 全フラグをログ出力

        Args:
            actor: アクター（デフォルト: pc）
        """
        return self.mod_invoke('debug_log_flags()', actor)

    def debug_log_quests(self, actor: Union[str, DramaActor] = None) -> 'DramaBuilder':
        """
        デバッグ: 全クエスト状態をログ出力

        Args:
            actor: アクター（デフォルト: pc）
        """
        return self.mod_invoke('debug_log_quests()', actor)

    def check_available_quests_for_npc(self, npc_id: str,
                                        actor: Union[str, DramaActor] = None) -> 'DramaBuilder':
        """
        NPCごとの利用可能クエストをチェックしフラグを設定

        設定されるフラグ:
        - sukutsu_available_quest_count: クエスト数
        - sukutsu_has_rank_up: ランクアップクエストの有無 (0/1)
        - sukutsu_has_character_event: キャラクターイベントの有無 (0/1)
        - sukutsu_has_sub_quest: サブクエストの有無 (0/1)
        - sukutsu_top_quest_id: 最優先クエストのハッシュ

        Args:
            npc_id: NPCのID（例: "sukutsu_arena_master"）
            actor: アクター（デフォルト: pc）
        """
        return self.mod_invoke(f'check_available_quests({npc_id})', actor)

    def check_quests_for_dispatch(self, flag_name: str, quest_ids: List[str]) -> 'DramaBuilder':
        """
        クエストディスパッチ用のチェック

        modInvokeでcheck_quests_for_dispatchコマンドを呼び出し、
        指定されたクエストIDリストの中で利用可能な最初のクエストのインデックスをフラグに設定する。

        設定されるフラグ値:
        - 0: 利用可能なクエストなし（fallback）
        - 1: リストの1番目のクエストが利用可能
        - 2: リストの2番目のクエストが利用可能
        - ...

        Args:
            flag_name: 設定するフラグ名
            quest_ids: クエストIDのリスト（優先度順）
        """
        # modInvoke形式: check_quests_for_dispatch(flagName, questId1, questId2, ...)
        args = [flag_name] + quest_ids
        param = f'check_quests_for_dispatch({", ".join(args)})'
        return self.action("modInvoke", param=param, actor="pc")

    def build(self) -> List[Dict[str, Any]]:
        """エントリーリストを返す（検証なし）"""
        return self.entries

    def _validate_drama_structure(self) -> List[str]:
        """
        ドラマ構造を検証し、警告メッセージのリストを返す

        検出ルール:
        1. 終端なしステップ: step後にend/jump/choiceがないステップ
        2. 孤立ステップ: 他から参照されていないステップ（mainを除く）
        3. 未定義ジャンプ: 存在しないラベルへのjump
        """
        import re
        warnings = []

        # ステップとその終端情報を収集
        steps = {}  # {step_name: {'has_terminator': bool, 'index': int}}
        jump_targets = set()  # 参照されているラベル
        current_step = None
        current_step_has_terminator = False
        current_step_has_fallback = False  # フォールバックジャンプがあるか

        for i, entry in enumerate(self.entries):
            if entry.get('step'):
                # 前のステップの終端チェック
                if current_step and not current_step_has_terminator:
                    # switch_on_flagパターン: 複数のmodInvoke + フォールバック（==0）がある場合は終端とみなす
                    if not current_step_has_fallback:
                        steps[current_step]['has_terminator'] = False

                current_step = entry['step']
                current_step_has_terminator = False
                current_step_has_fallback = False
                steps[current_step] = {'has_terminator': True, 'index': i}

            # 終端アクションのチェック
            action = entry.get('action')
            jump = entry.get('jump')
            param = entry.get('param', '')

            if action in ('end', 'choice', 'cancel'):
                current_step_has_terminator = True
            if jump:
                current_step_has_terminator = True
                jump_targets.add(jump)

            # modInvoke内のif_flagジャンプターゲットを抽出
            if action in ('modInvoke', 'invoke*') and param:
                # if_flag(flag, ==value, target) パターンを検出
                match = re.search(r'if_flag\([^,]+,\s*[^,]+,\s*([^)]+)\)', param)
                if match:
                    target = match.group(1).strip()
                    jump_targets.add(target)
                    # ==0のフォールバックがある場合、このステップは終端とみなす
                    if '==0' in param:
                        current_step_has_fallback = True

                # check_quest_available(quest_id, target) パターンを検出
                match = re.search(r'check_quest_available\([^,]+,\s*([^)]+)\)', param)
                if match:
                    target = match.group(1).strip()
                    jump_targets.add(target)

        # 最後のステップのチェック
        if current_step and not current_step_has_terminator and not current_step_has_fallback:
            steps[current_step]['has_terminator'] = False

        # 警告生成
        for step_name, info in steps.items():
            if not info['has_terminator']:
                warnings.append(f"Step '{step_name}' has no terminator (end/jump/choice)")

        # 孤立ステップのチェック（main以外）
        for step_name in steps:
            if step_name != 'main' and step_name not in jump_targets:
                # choice_blockで生成される_reactステップは除外
                if '_react_' not in step_name:
                    warnings.append(f"Step '{step_name}' is never referenced (orphan)")

        # 未定義ジャンプのチェック
        defined_steps = set(steps.keys())
        # 組み込みステップは除外
        builtin_steps = {'_trade', '_buy', '_joinParty', '_leaveParty', '_train',
                        '_heal', '_investShop', '_sellFame', '_copyItem', '_give',
                        '_whore', '_tail', '_suck', '_bout', '_rumor', '_news'}
        for target in jump_targets:
            if target not in defined_steps and target not in builtin_steps:
                warnings.append(f"Jump target '{target}' is not defined")

        return warnings

    def save(self, filepath: str, sheet_name: str = "main") -> None:
        """
        CWL準拠のExcelファイルとして保存。

        Args:
            filepath: 出力ファイルパス
            sheet_name: シート名（キャラクターIDなどを推奨）
        """
        # バリデーション実行
        warnings = self._validate_drama_structure()
        for w in warnings:
            print(f"[DramaBuilder] WARNING: {w}")

        os.makedirs(os.path.dirname(filepath), exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Row 1: Headers
        for col, header in enumerate(HEADERS, 1):
            ws.cell(row=1, column=col, value=header)

        # Row 2-5: 空行（CWL形式に準拠）

        # Row 6+: Data
        row = 6
        for entry in self.entries:
            for col, header in enumerate(HEADERS, 1):
                value = entry.get(header)
                if value is not None:
                    ws.cell(row=row, column=col, value=value)
            row += 1

        wb.save(filepath)
        print(f"Created: {filepath} (sheet: {sheet_name})")
