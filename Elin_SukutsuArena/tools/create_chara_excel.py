import openpyxl
import os
import csv

# パス設定
TOOLS_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(TOOLS_DIR)
SamplePath = r'c:\Users\tishi\programming\elin_modding\CWL_AddLocation_Example\LangMod\EN\SourceSSS.xlsx'
OUTPUT_EN_TSV = os.path.join(PROJECT_ROOT, 'LangMod', 'EN', 'Chara.tsv')
OUTPUT_JP_TSV = os.path.join(PROJECT_ROOT, 'LangMod', 'JP', 'Chara.tsv')

print(f'Reading sample file from: {SamplePath}')

sample_wb = openpyxl.load_workbook(SamplePath)
sample_ws = sample_wb['Chara']

# Row 1-3 (Header, Type, Default) Copy
header_map = {}
cols = sample_ws.max_column

rows = []
for r in range(1, 4):
    row_data = []
    for c in range(1, cols + 1):
        cell_val = sample_ws.cell(row=r, column=c).value
        if r == 1:
            header_map[c-1] = cell_val
        row_data.append(cell_val if cell_val is not None else "")
    rows.append(row_data)

# Helper
def create_npc_row(npc_def):
    row = [""] * cols
    for k, v in npc_def.items():
        found = False
        for col_idx, col_name in header_map.items():
            if col_name == k:
                row[col_idx] = v
                found = True
                break
        if not found:
            print(f"WARNING: Field '{k}' not found in headers!")
    return row

npcs = []

# ===== NPC定義 =====
# bio format: 性別(m/f) / バージョン(固定ID) / 身長 / 体重 / 性格 | 髪色 | 肌色
# _idRenderData: @chara (PCCシステムを使用、Texture/ID.pngが自動ロードされる)
# tiles: バニラのフォールバックタイルID (カスタム画像が見つからない時に使用)
# idText: テキストID (同じIDのテキストファイルが参照される)

# 1. アリーナ受付嬢 (サキュバス / 女性 / 固定外見)
npcs.append({
    'id': 'sukutsu_receptionist',
    'name_JP': 'アリーナ受付嬢',
    'name': 'Arena Receptionist',
    'aka_JP': '魅惑の案内人',
    'aka': 'Charming Guide',
    'race': 'succubus',
    'job': 'shopkeeper',
    'LV': 50,
    'hostility': 'Friend',
    'tiles': 807,  # フォールバック用バニラタイル
    '_idRenderData': '@chara',  # PCCシステム使用
    'bio': 'f/1001/165/52/sexy',  # 女性/固定ID/身長/体重/性格
    'idText': 'sukutsu_receptionist',
    'tag': 'neutral,addStock',  # 商人タグ
    'trait': 'Merchant',
    'quality': 4,
    'chance': 0,  # ランダムスポーン無効
})

# 2. アリーナマスター (人間 / 男性 / 戦士)
npcs.append({
    'id': 'sukutsu_arena_master',
    'name_JP': 'アリーナマスター',
    'name': 'Arena Master',
    'aka_JP': '百戦の覇者',
    'aka': 'Champion of Hundred Battles',
    'race': 'human',
    'job': 'warrior',
    'LV': 70,
    'hostility': 'Friend',
    'tiles': 807,
    '_idRenderData': '@chara',
    'bio': 'm/1002/185/90/stern',  # 男性/固定ID/身長/体重/性格
    'idText': 'sukutsu_arena_master',
    'tag': 'neutral,addDrama_drama_sukutsu_arena_master',  # CWLドラマスクリプト使用
    'quality': 4,
    'chance': 0,
})

# 3. グランドマスター (ドラゴン / 男性 / チャンピオン)
npcs.append({
    'id': 'sukutsu_grand_master',
    'name_JP': 'グランドマスター',
    'name': 'Grand Master',
    'aka_JP': '竜鱗の王者',
    'aka': 'Dragon Scale Champion',
    'race': 'dragon',
    'job': 'warrior',
    'LV': 100,
    'hostility': 'Friend',
    'tiles': 807,
    '_idRenderData': '@chara',
    'bio': 'm/1003/210/150/proud',  # 男性/固定ID/身長/体重/性格
    'idText': 'sukutsu_grand_master',
    'tag': 'neutral',
    'quality': 5,
    'chance': 0,
})

# 4. 怪しい商人 (ミュータント / 性別不明 / 商人)
npcs.append({
    'id': 'sukutsu_shady_merchant',
    'name_JP': '怪しい商人',
    'name': 'Shady Merchant',
    'aka_JP': '闇市の支配者',
    'aka': 'Lord of Black Market',
    'race': 'mutant',
    'job': 'merchant',
    'LV': 60,
    'hostility': 'Friend',
    'tiles': 807,
    '_idRenderData': '@chara',
    'bio': 'm/1004/170/65/sly',  # 男性/固定ID/身長/体重/性格
    'idText': 'sukutsu_shady_merchant',
    'tag': 'neutral,addStock',
    'trait': 'Merchant',
    'quality': 4,
    'chance': 0,
})

for npc in npcs:
    rows.append(create_npc_row(npc))

def write_tsv(path, row_data):
    with open(path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f, delimiter='\t', quoting=csv.QUOTE_MINIMAL)
        writer.writerows(row_data)
    print(f'Created TSV: {path}')

write_tsv(OUTPUT_EN_TSV, rows)
write_tsv(OUTPUT_JP_TSV, rows)
