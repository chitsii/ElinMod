import openpyxl
import os
import csv

# パス設定
BUILDER_DIR = os.path.dirname(os.path.abspath(__file__))
TOOLS_DIR = os.path.dirname(BUILDER_DIR)
PROJECT_ROOT = os.path.dirname(TOOLS_DIR)
SAMPLE_PATH = r'c:\Users\tishi\programming\elin_modding\CWL_AddLocation_Example\LangMod\EN\SourceSSS.xlsx'
OUTPUT_EN_TSV = os.path.join(PROJECT_ROOT, 'LangMod', 'EN', 'Zone.tsv')
OUTPUT_JP_TSV = os.path.join(PROJECT_ROOT, 'LangMod', 'JP', 'Zone.tsv')

print(f'Reading sample file from: {SAMPLE_PATH}')

# サンプル読み込み
sample_wb = openpyxl.load_workbook(SAMPLE_PATH)
sample_ws = sample_wb['Zone']

# TSVデータ構築
rows = []

# Row 1-3 (Header, Type, Default) Copy
for r in range(1, 4):
    row_data = []
    for c in range(1, sample_ws.max_column + 1):
        cell_val = sample_ws.cell(row=r, column=c).value
        # Noneを空文字に変換しないと "None" と書かれる恐れがあるが、
        # CWLのExcelパーサーの期待による。通常は空文字で良い。
        row_data.append(cell_val if cell_val is not None else "")
    rows.append(row_data)

# Row 4 (Data) Construction
# まず空のリストを作成（カラム数分）
data_row = [""] * sample_ws.max_column

# 必要なデータを埋める (Indexは0始まりなので カラム番号-1)
# Headerマップを作ると安全だが、ここでは固定インデックスでいく（前回と同じ）
# A=0, B=1, ...
data_row[0] = 'sukutsu_arena' # A: id
data_row[1] = 'ntyris'        # B: parent
data_row[2] = '巣窟アリーナ'   # C: name_JP
data_row[3] = 'Sukutsu Arena' # D: name
data_row[4] = 'Elin_SukutsuArena.Zone_SukutsuArena' # E: type
data_row[5] = 50  # F: LV
data_row[6] = 100 # G: chance
# H: faction (skipped/default)
data_row[8] = 100 # I: value
# J: idProfile
data_row[10] = 'sukutsu_arena' # K: idFile
data_row[11] = 'Plain' # L: idBiome
# M: idGen
# N: idPlaylist
data_row[14] = 'addMap,light' # O: tag
# P: cost
# Q: dev
data_row[17] = 'default' # R: image (User requested 'default')
data_row[18] = '34,-24,323' # S: pos
# T: questTag
data_row[20] = '熱狂と興奮が渦巻く、地下闘技場への入り口。' # U: textFlavor_JP
data_row[21] = 'The entrance to the underground arena.' # V: textFlavor

rows.append(data_row)

# 新規: field_fine エントリ（通常バトルゾーン用）
data_row_fine = [""] * sample_ws.max_column
data_row_fine[0] = 'field_fine'                               # A: id
data_row_fine[1] = ''                                         # B: parent
data_row_fine[2] = '闘技場フィールド'                          # C: name_JP
data_row_fine[3] = 'Arena Field'                              # D: name
data_row_fine[4] = 'Elin_SukutsuArena.Zone_FieldFine'        # E: type
data_row_fine[5] = 1                                          # F: LV
data_row_fine[6] = 100                                        # G: chance
data_row_fine[10] = 'chitsii_battle_field_fine'              # K: idFile（カスタムマップ）
data_row_fine[11] = 'Plain'                                   # L: idBiome
data_row_fine[14] = 'addMap'                                  # O: tag (light不要)
rows.append(data_row_fine)

# 新規: field_snow エントリ（雪原バトルゾーン用）
data_row_snow = [""] * sample_ws.max_column
data_row_snow[0] = 'field_snow'                              # A: id
data_row_snow[1] = ''                                         # B: parent
data_row_snow[2] = '雪原フィールド'                            # C: name_JP
data_row_snow[3] = 'Snow Field'                               # D: name
data_row_snow[4] = 'Elin_SukutsuArena.Zone_FieldSnow'        # E: type
data_row_snow[5] = 1                                          # F: LV
data_row_snow[6] = 100                                        # G: chance
data_row_snow[10] = 'chitsii_battle_field_snow'              # K: idFile（カスタムマップ）
data_row_snow[11] = 'Snow'                                    # L: idBiome
data_row_snow[14] = 'addMap'                                  # O: tag (light不要)
rows.append(data_row_snow)

# TSV書き出し関数
def write_tsv(path, row_data):
    with open(path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f, delimiter='\t', quoting=csv.QUOTE_MINIMAL)
        writer.writerows(row_data)
    print(f'Created TSV: {path}')

write_tsv(OUTPUT_EN_TSV, rows)
write_tsv(OUTPUT_JP_TSV, rows)
