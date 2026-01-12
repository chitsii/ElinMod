"""
SourceStat.xlsx 生成スクリプト

アスタロトの権能（カスタムCondition）をゲームに登録するためのExcelファイルを生成する。
CWLはこのファイルを読み込み、SourceStatテーブルにエントリを追加する。
"""
import openpyxl
import os

# パス設定
BUILDER_DIR = os.path.dirname(os.path.abspath(__file__))
TOOLS_DIR = os.path.dirname(BUILDER_DIR)
PROJECT_ROOT = os.path.dirname(TOOLS_DIR)

OUTPUT_JP = os.path.join(PROJECT_ROOT, 'LangMod', 'JP', 'SourceStat.xlsx')
OUTPUT_EN = os.path.join(PROJECT_ROOT, 'LangMod', 'EN', 'SourceStat.xlsx')

# SourceStat のヘッダー列（SourceStat.cs より）
# CWLのNamedImportを使用するため、列名でマッピングされる
HEADERS = [
    'id',           # ユニークID (int)
    'alias',        # クラス名 (string)
    'name_JP',      # 日本語名 (string)
    'name',         # 英語名 (string)
    'type',         # C#完全修飾クラス名 (string)
    'group',        # グループ (string)
    'curse',        # 呪い (string)
    'duration',     # 持続時間 (string)
    'hexPower',     # 呪い力 (int)
    'negate',       # 打消し対象 (string[])
    'defenseAttb',  # 防御属性 (string[])
    'resistance',   # 耐性 (string[])
    'gainRes',      # 耐性獲得 (int)
    'elements',     # 要素 (string[])
    'nullify',      # 無効化 (string[])
    'tag',          # タグ (string[])
    'phase',        # フェーズ閾値 (int[])
    'colors',       # 色 (string)
    'element',      # 元素 (string)
    'effect',       # エフェクト (string[])
    'strPhase_JP',  # フェーズ文字列JP (string[])
    'strPhase',     # フェーズ文字列EN (string[])
    'textPhase_JP', # フェーズテキストJP (string)
    'textPhase',    # フェーズテキストEN (string)
    'textEnd_JP',   # 終了テキストJP (string)
    'textEnd',      # 終了テキストEN (string)
    'textPhase2_JP',# フェーズ2テキストJP (string)
    'textPhase2',   # フェーズ2テキストEN (string)
    'gradient',     # グラデーション (string)
    'invert',       # 反転 (bool)
    'detail_JP',    # 詳細JP (string)
    'detail',       # 詳細EN (string)
]

# アスタロトの権能（カスタムCondition）定義
CONDITIONS = [
    {
        'id': 10001,
        'alias': 'ConAstarothTyranny',
        'name_JP': '時の独裁',
        'name': 'Tyranny of Time',
        'type': 'Elin_SukutsuArena.Conditions.ConAstarothTyranny',
        'detail_JP': 'アスタロトの権能により、速度が大幅に低下している。',
        'detail': 'Your speed is drastically reduced by Astaroth\'s power.',
    },
    {
        'id': 10002,
        'alias': 'ConAstarothDenial',
        'name_JP': '因果の拒絶',
        'name': 'Denial of Causality',
        'type': 'Elin_SukutsuArena.Conditions.ConAstarothDenial',
        'detail_JP': 'アスタロトの権能により、筋力が大幅に低下している。',
        'detail': 'Your strength is drastically reduced by Astaroth\'s power.',
    },
    {
        'id': 10003,
        'alias': 'ConAstarothDeletion',
        'name_JP': '終焉の削除命令',
        'name': 'Deletion Command of the End',
        'type': 'Elin_SukutsuArena.Conditions.ConAstarothDeletion',
        'detail_JP': 'アスタロトの権能により、魔力が消し去られている。MPが0になる。',
        'detail': 'Your magical power is erased by Astaroth\'s power. MP is reduced to 0.',
    },
]


def create_condition_row(condition_def):
    """Conditionの行データを作成"""
    row = [''] * len(HEADERS)
    for key, value in condition_def.items():
        if key in HEADERS:
            idx = HEADERS.index(key)
            row[idx] = value
    return row


def create_excel(output_path, lang='JP'):
    """SourceStat.xlsx を生成"""
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Stat'  # CWLが期待する表名

    # Row 1: ヘッダー
    for col, header in enumerate(HEADERS, 1):
        ws.cell(row=1, column=col, value=header)

    # Row 2-3: 型定義とデフォルト値（空行でOK、CWLがスキップ）
    # CWLはRow 6以降をデータとして読み込む

    # Row 6+: データ
    row_num = 6
    for condition in CONDITIONS:
        row_data = create_condition_row(condition)
        for col, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col, value=value)
        row_num += 1

    wb.save(output_path)
    print(f'Created: {output_path}')


def main():
    print('Generating SourceStat.xlsx for custom Conditions...')

    # JP版とEN版を生成
    create_excel(OUTPUT_JP, 'JP')
    create_excel(OUTPUT_EN, 'EN')

    print('Done!')


if __name__ == '__main__':
    main()
