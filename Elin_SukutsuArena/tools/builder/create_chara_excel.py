import openpyxl
import os
import csv
import argparse

# コマンドライン引数の解析
parser = argparse.ArgumentParser(description='Create Chara Excel for Elin_SukutsuArena')
parser.add_argument('--debug', action='store_true', help='Debug mode: set all boss LV to 1')
args = parser.parse_args()

DEBUG_MODE = args.debug
if DEBUG_MODE:
    print("[DEBUG MODE] All boss characters will have LV=1")

# パス設定
# パス設定
BUILDER_DIR = os.path.dirname(os.path.abspath(__file__))
TOOLS_DIR = os.path.dirname(BUILDER_DIR)
PROJECT_ROOT = os.path.dirname(TOOLS_DIR)
SamplePath = r'c:\Users\tishi\programming\elin_modding\CWL_AddLocation_Example\LangMod\EN\SourceSSS.xlsx'
OUTPUT_EN_TSV = os.path.join(PROJECT_ROOT, 'LangMod', 'EN', 'Chara.tsv')
OUTPUT_JP_TSV = os.path.join(PROJECT_ROOT, 'LangMod', 'JP', 'Chara.tsv')

print(f'Reading sample file from: {SamplePath}')

sample_wb = openpyxl.load_workbook(SamplePath)
sample_ws = sample_wb['Chara']

# Row 1-3 (Header, Type, Default) Copy
header_map = {}
cols = sample_ws.max_column

rows = []
for r in range(1, 4):
    row_data = []
    for c in range(1, cols + 1):
        cell_val = sample_ws.cell(row=r, column=c).value
        # 1行目ならヘッダーマップ作成
        if r == 1:
            header_map[c-1] = cell_val
        row_data.append(cell_val if cell_val is not None else "")
    rows.append(row_data)

# 'Author' カラムが存在しない場合、強制的に追加
if 'Author' not in header_map.values():
    print("Adding missing 'Author' column...")
    new_col_idx = len(header_map)
    header_map[new_col_idx] = 'Author'

    # Header行に追加
    rows[0].append('Author')
    # Type行, Default行に追加 (空文字)
    rows[1].append('')
    rows[2].append('')

    # colsを更新
    cols += 1

# 'portrait' カラムが存在しない場合、強制的に追加
if 'portrait' not in header_map.values():
    print("Adding missing 'portrait' column...")
    new_col_idx = len(header_map)
    header_map[new_col_idx] = 'portrait'

    # Header行に追加
    rows[0].append('portrait')
    # Type行, Default行に追加 (空文字)
    rows[1].append('')
    rows[2].append('')

    # colsを更新
    cols += 1

# Helper
def create_npc_row(npc_def):
    row = [""] * cols  # 更新されたcolsを使用
    for k, v in npc_def.items():
        found = False
        for col_idx, col_name in header_map.items():
            if col_name == k:
                row[col_idx] = v
                found = True
                break
        if not found:
            # authorなどはここで処理済みのはずだが、もしヘッダーになければ警告
             print(f"WARNING: Field '{k}' not found in headers!")
    return row

npcs = []

# ===== NPC定義 =====
# bio format: 性別(m/f) / バージョン(固定ID) / 身長 / 体重 / 性格 | 髪色 | 肌色
# _idRenderData: @chara (PCCシステムを使用、Texture/ID.pngが自動ロードされる)
# tiles: バニラのフォールバックタイルID (カスタム画像が見つからない時に使用)
# idText: テキストID (同じIDのテキストファイルが参照される)
#
# CWL タグ仕様:
# - addZone_ゾーンID: 指定ゾーンにキャラクターを生成
# - addFlag_StayHomeZone: ランダム移動を無効化（初期ゾーンに留まる）
# - addDrama_テーブル名: ドラマシートをリンク
# - humanSpeak: 人間らしい会話表示（括弧なし）
# - addStock: 商人の在庫を追加

ZONE_ID = 'sukutsu_arena'  # カスタムゾーンID

# キャラクターレンダリング用デフォルト設定
# _idRenderData: 'chara' = バニラキャラスプライトシートを使用
# tiles: タイルID（バニラキャラのスプライト位置）

# 1. リリィ (サキュバス / 女性 / 受付嬢)
npcs.append({
    'id': 'sukutsu_receptionist',
    'Author': 'tishi.elin.sukutsu_arena',
    'name_JP': 'リリシエル',
    'name': 'Lilithiel',
    'aka_JP': '登録官',
    'aka': 'The Overseer',
    'race': 'succubus',
    'job': 'shopkeeper',
    '_idRenderData': '@chara',
    'tiles': 340,  # succubus female
    'LV': 500,
    'hostility': 'Friend',
    'bio': 'f/1001/165/52/sexy',
    'idText': 'sukutsu_receptionist',
    # CWL タグ: ゾーン生成、ランダム移動無効、商人在庫、ドラマリンク、人間らしい会話
    'tag': f'neutral,addZone_{ZONE_ID},addFlag_StayHomeZone,addStock_sukutsu_receptionist,addDrama_drama_sukutsu_receptionist,humanSpeak',
    'trait': 'SukutsuMerchant',  # メインクエスト完了前はペット化不可
    'quality': 4,
    'chance': 0,
})

# 2. バルガス (人間 / 男性 / アリーナマスター)
# Note: 敵バージョン（訓練・全盛期）と共通のベース定義を使用
_sukutsu_balgas_base = {
    'id': 'sukutsu_arena_master',
    'Author': 'tishi.elin.sukutsu_arena',
    'name_JP': 'バルガス',
    'name': 'Vargus',
    'aka_JP': 'アリーナマスター',
    'aka': 'Champion of Hundred Battles',
    'race': 'juere',
    'job': 'warrior',
    '_idRenderData': '@chara',
    'tiles': 0,  # human male warrior
    'LV': 800,
    'hostility': 'Friend',
    'bio': 'm/1002/185/90/stern',
    'idText': 'sukutsu_arena_master',  # 共通テキストID
    # CWL タグ: ゾーン生成、ランダム移動無効、ドラマリンク、人間らしい会話
    'tag': f'neutral,addZone_{ZONE_ID},addFlag_StayHomeZone,addDrama_drama_sukutsu_arena_master,humanSpeak',
    'trait': 'SukutsuNPC',  # メインクエスト完了前はペット化不可
    'quality': 4,
    'chance': 0,
}
npcs.append(_sukutsu_balgas_base)

# 3. アスタロト (ドラゴン / 男性 / アリーナ創設者 / 最終ボス)
# lore: イルヴァの神々と同格の竜神、Lv.50000
# 滅びた次元「カラドリウス」の唯一の生存者、アリーナ創設者
# 3,000年間、観客の「注目」を力に変換してきた
# Note: アリーナではNPCとして配置、ボス戦時は新規スポーン
npcs.append({
    'id': 'sukutsu_astaroth',
    'Author': 'tishi.elin.sukutsu_arena',
    'name_JP': 'アスタロト',
    'name': 'Astaroth',
    'aka_JP': 'うつろいし竜神',
    'aka': 'The Hollow Dragon God',
    'race': 'dragon',
    'job': 'warrior',
    '_idRenderData': '@chara',
    'tiles': 168,  # dragon
    'LV': 50000,  # 最終ボス級
    'hostility': 'Friend',  # アリーナではNPCとして配置
    'bio': 'm/1003/350/500/proud',
    'idText': 'sukutsu_astaroth',
    'portrait': 'UN_sukutsu_astaroth',  # カスタムポートレート
    # 能力: 全属性攻撃、12部位、ほぼ全属性耐性（聖のみ弱点）、大幅強化
    'mainElement': 'Void',
    'elements': 'featElder/1,featBodyParts/12,resVoid/20,resChaos/20,resNether/20,resMagic/20,resDarkness/20,resFire/20,resNerve/20,resMind/20,resHoly/20,resCut/20,resImpact/20,featBoost/1,featBloodBond/1',
    'actCombat': 'breathe_Void/35,breathe_Chaos/25,breathe_Nether/20,hand_Magic/30,SpGravity/10,SpBane/5,ActGazeInsane/15,ActGazeMutation/10,ActCurse/10,ActBurnMana/15,SpHeal/5,SpSummonDragon/5,SpEarthquake/5,SpShutterHex/5,SpSpeedDown/5,SpSilence/5,SpWeakness/5,SpNightmare/5,SpSummonTentacle/5,SpGate/5,ActTouchDrown/5,ActNeckHunt/5',
    # CWL タグ: ゾーン生成、ランダム移動無効
    'tag': f'boss,undead,addZone_{ZONE_ID},addFlag_StayHomeZone',
    'trait': 'SukutsuNPC',  # メインクエスト完了前はペット化不可
    'quality': 5,
    'chance': 0,
})

# 4. 怪しい商人 (ミュータント / 性別不明 / 商人)
npcs.append({
    'id': 'sukutsu_shady_merchant',
    'Author': 'tishi.elin.sukutsu_arena',
    'name_JP': 'ゼク',
    'name': 'Ezekiel',
    'aka_JP': '剥製師',
    'aka': 'The Taxidermist',
    'race': 'mutant',
    'job': 'merchant',
    'LV': 666,
    'hostility': 'Friend',
    'tiles': 807,
    '_idRenderData': '@chara',
    'bio': 'm/1004/170/65/sly',
    'idText': 'sukutsu_shady_merchant',
    # CWL タグ: ゾーン生成、ランダム移動無効、商人在庫、ドラマリンク、人間らしい会話
    'tag': f'neutral,addZone_{ZONE_ID},addFlag_StayHomeZone,addStock_sukutsu_shady_merchant,addDrama_drama_sukutsu_shady_merchant,humanSpeak',
    'trait': 'SukutsuMerchant',  # メインクエスト完了前はペット化不可
    'quality': 4,
    'chance': 0,
})

# 5. デバッグマスター (開発テスト用 / 各バトルに直接アクセス可能)
npcs.append({
    'id': 'sukutsu_debug_master',
    'Author': 'tishi.elin.sukutsu_arena',
    'name_JP': '観測者',
    'name': 'Observer',
    'aka_JP': '次元の記録者',
    'aka': 'Dimensional Recorder',
    'race': 'spirit',
    'job': 'wizard',
    '_idRenderData': '@chara',
    'tiles': 478,  # spirit/wisp
    'LV': 1,
    'hostility': 'Friend',
    'bio': 'n/1005/160/0/mysterious',
    'idText': 'sukutsu_debug_master',
    # CWL タグ: ゾーン生成、ランダム移動無効、ドラマリンク、人間らしい会話
    'tag': f'neutral,addZone_{ZONE_ID},addFlag_StayHomeZone,addDrama_drama_debug_menu,humanSpeak',
    'quality': 5,
    'chance': 0,
})

# ===== ストーリーバトル用敵キャラクター =====
# これらはアリーナゾーンには配置せず、バトル時に動的にスポーンさせる

# 6. 虚空のウーズ (Rank G昇格試験 / 混沌ブレススライム群)
# lore: 次元の狭間で変異したスライム、混沌ブレスを吐く
npcs.append({
    'id': 'sukutsu_void_ooze',
    'Author': 'tishi.elin.sukutsu_arena',
    'name_JP': '虚空のウーズ',
    'name': 'Void Ooze',
    'aka_JP': '混沌の落とし子',
    'aka': 'Child of Chaos',
    'race': 'slime',
    'job': 'warrior',
    '_idRenderData': '@chara',
    'tiles': 296,  # slime
    'LV': 15,  # Tier1初戦: Lv.1-100
    'hostility': 'Enemy',
    'bio': 'n/2001/50/30/mindless',
    'idText': 'sukutsu_void_ooze',
    # 能力: 混沌ブレス、混沌免疫、肉クッション、窃盗系
    'mainElement': 'Chaos',
    'elements': 'resChaos/100,featMeatCushion/4',
    'actCombat': 'breathe_Chaos/30,ActStealFood/15,ActStealMoney/15,ActDraw/10,ActTouchDrown/10',
    'tag': 'boss',
    'quality': 3,
    'chance': 0,
})

# 7. 凍土の猟犬 (Rank F昇格試験 / 古代種・氷・透明魔犬)
# lore: 古代種の魔犬、氷ブレスと透明化で獲物を狩る
npcs.append({
    'id': 'sukutsu_frost_hound',
    'Author': 'tishi.elin.sukutsu_arena',
    'name_JP': '凍土の猟犬',
    'name': 'Frostfang Hound',
    'aka_JP': '凍える牙',
    'aka': 'Freezing Fang',
    'race': 'hound',
    'job': 'warrior',
    '_idRenderData': '@chara',
    'tiles': 254,  # hound
    'LV': 45,  # Tier1中盤: Lv.1-100
    'hostility': 'Enemy',
    'bio': 'n/2002/120/80/predator',
    'idText': 'sukutsu_frost_hound',
    # 能力: 古代種、透明化、氷ブレス、鈍足魔法、氷免疫
    'mainElement': 'Cold',
    'elements': 'invisibility/1,featElder/1,resCold/100',  # 透明化、古代種、氷免疫
    'actCombat': 'breathe_Cold/40,SpSpeedDown/20',  # 氷ブレス40%、鈍足20%
    'tag': 'boss',
    'quality': 4,
    'chance': 0,
})

# 8. 訓練用バルガス (バルガス訓練クエスト / 手加減バージョン)
# lore: 「お前の足を一歩でも動かしてみせろ」
# 全力ではないが、それでも歴戦の覇者
# Note: _sukutsu_balgas_baseをコピーし、必要な部分のみ上書き
_sukutsu_balgas_training = _sukutsu_balgas_base.copy()
_sukutsu_balgas_training.update({
    'id': 'sukutsu_balgas_training',
    'LV': 200,  # 手加減バージョン
    'hostility': 'Enemy',
    # 能力: 物理特化（手加減）
    'mainElement': 'Cut',
    'elements': 'featElder/1,resCut/60,resImpact/60',
    'actCombat': '',
    'tag': 'boss',  # addZoneなし（バトル専用）
    'trait': '',  # トレイト不要（バトル専用）
})
npcs.append(_sukutsu_balgas_training)

# 9. 全盛期バルガス (Rank S昇格試験 / 若返った最強の戦士)
# lore: 68歳の現在はLv.85、全盛期（30代）はLv.120相当
# 「若返りの薬」使用条件は「真に命を賭けた戦いの場」でのみ
# 純粋な「武の極致」。魔法もブレスも使わず、剣と体術のみで全てを圧倒する
# Note: _sukutsu_balgas_baseをコピーし、必要な部分のみ上書き
_sukutsu_balgas_prime = _sukutsu_balgas_base.copy()
_sukutsu_balgas_prime.update({
    'id': 'sukutsu_balgas_prime',
    'name_JP': '全盛期のバルガス',
    'name': 'Balgas at His Prime',
    'aka_JP': '鉄血の覇者',
    'aka': 'Iron-Blooded Champion',
    'LV': 8000,  # Tier3: Lv.1,000-10,000
    'hostility': 'Enemy',
    'bio': 'm/1002/188/95/stern',
    # 能力: 物理特化、8部位連撃、高物理耐性、魔法無効、ヴォーパル、連撃系
    'mainElement': 'Cut',
    'elements': 'featElder/1,featBodyParts/8,resCut/80,resImpact/80,resFire/40,resCold/40,resLightning/40,resMind/60,resNerve/60,antiMagic/60,vopal/60,mod_flurry/65,mod_chaser/30,mod_cleave/10,redirect_blaser/23,counter/20',
    'actCombat': 'breathe_Cut/40,breathe_Impact/30,SpHero/10,ActRush/25,ActBash/20',
    'tag': 'boss',  # addZoneなし（バトル専用）
    'trait': '',  # トレイト不要（バトル専用）
    'quality': 5,
})
npcs.append(_sukutsu_balgas_prime)

# 10. カインの亡霊 (Rank E昇格試験 / バルガスの元副官)
# lore: 次元の狭間で変異したカオスシェイプ
# 記憶を失い、戦闘本能だけが残った異形の存在
npcs.append({
    'id': 'sukutsu_kain_ghost',
    'Author': 'tishi.elin.sukutsu_arena',
    'name_JP': '錆びついた英雄カイン',
    'name': 'Rusted Hero Cain',
    'aka_JP': '忘れられた副団長',
    'aka': 'Forgotten Vice-Captain',
    'race': 'wraith',
    'job': 'warrior',
    '_idRenderData': '@chara',
    'tiles': 458,  # ghost (fallback)
    'LV': 90,  # Tier1最終: Lv.1-100
    'hostility': 'Enemy',
    'bio': 'm/1006/180/0/melancholic',
    'idText': 'sukutsu_kain_ghost',
    # 能力: 多属性攻撃、追加部位、アンデッド、影召喚
    'elements': 'featBodyParts/5,featUndead/1',
    'actCombat': 'breathe_Acid/25,breathe_Cut/25,hand_Nether/30,SpSummonShadow/15',
    'tag': 'boss,undead',
    'quality': 4,
    'chance': 0,
})

# 9. ヌル (Rank B昇格試験 / 暗殺人形 / 人造生命体)
# lore: 「神の孵化場」計画の失敗作、アリーナの「清掃係」
# 透明化 + 分裂能力で数で圧倒する暗殺者
# Note: アリーナではNPCとして配置、ボス戦時は敵バージョン(sukutsu_null_enemy)を使用
_sukutsu_null_base = {
    'id': 'sukutsu_null',
    'Author': 'tishi.elin.sukutsu_arena',
    'name_JP': 'Nul',
    'name': 'Nul',
    'aka_JP': '虚無の処刑人',
    'aka': 'Void Executioner',
    'race': 'machine',
    'job': 'thief',
    '_idRenderData': '@chara',
    'tiles': 536,  # machine/robot
    'LV': 2000,
    'hostility': 'Friend',  # アリーナではNPCとして配置
    'bio': 'f/1007/165/45/princess',
    'idText': 'sukutsu_null',
    'portrait': 'UN_sukutsu_null',  # カスタムポートレート
    # 能力: 大幅強化版
    'mainElement': 'Void',
    'elements': 'invisibility/1,featSplit/1,featElder/1,resVoid/80,resNether/60,resMagic/40,resNerve/100,resMind/100,featGolem/1,featReboot/1,featBoost/1,featEarthStrength/1,featRapidArrow/3,featGeneSlot/10,featMiscreation/1,featMetal/120,featManaMeat/1,featRoran/1,evasionPerfect/60',
    'actCombat': 'hand_Void/40,SpInvisibility/30,SpSilence/15,ActGazeInsane/15,ActRush/10,ActInsult/10,SpEarthquake/10,SpSeeInvisible/10,arrow_Void/20',
    # アリーナに配置してglobalCharasに登録
    'tag': f'boss,addZone_{ZONE_ID},addFlag_StayHomeZone',
    'trait': 'SukutsuNPC',  # メインクエスト完了前はペット化不可
    'quality': 4,
    'chance': 0,
}
npcs.append(_sukutsu_null_base)

# 9b. Nul（敵バージョン / バトル用）
# sukutsu_null_baseをコピーし、必要な部分のみ上書き
# 分身（featSplit）もEnemyとして生成される
_sukutsu_null_enemy = _sukutsu_null_base.copy()
_sukutsu_null_enemy.update({
    'id': 'sukutsu_null_enemy',
    'hostility': 'Enemy',  # バトル用: 敵として生成
    'tag': 'boss',  # addZoneなし（バトル専用、アリーナに配置しない）
    'trait': '',  # トレイト不要（バトル専用）
})
npcs.append(_sukutsu_null_enemy)

# 10. 影のドッペルゲンガー (Rank A昇格試験 / プレイヤーの影)
# lore: 観客の「注目」がプレイヤーの影から生み出した存在
# プレイヤーの全てを知り、全てを模倣する
# Note: 影の自己はC#コードでPCの外見をコピーするため、tilesは仮の値
npcs.append({
    'id': 'sukutsu_shadow_self',
    'Author': 'tishi.elin.sukutsu_arena',
    'name_JP': '影の自己',
    'name': 'Shadow Self',
    'aka_JP': '映し身',
    'aka': 'Mirror Image',
    'race': 'phantom',
    'job': 'warrior',
    '_idRenderData': '@chara',
    'tiles': 460,
    'LV': 2500,  # Tier3前半: Lv.1,000-10,000
    'hostility': 'Enemy',
    'bio': 'n/1008/170/60/sinister',
    'idText': 'sukutsu_shadow_self',
    # 試練ボス: プレイヤーの影、デーモン
    'elements': 'featDemon/1',
    'actCombat': 'arrow_Darkness/25,SpHeal/15,SpTeleport/15,ActSwarm/20,ActBladeStorm/20',
    'tag': 'boss,undead',
    'quality': 4,
    'chance': 0,
})

# 11. グリード (Rank D昇格試験 / 観客の代弁者)
# lore: かつてCランク「闘技場の鴉」の一員だった剣士
# 観客の「注目」に魅せられ、自ら観客の力を体に取り込もうとした
# 結果、肉体は半ば「観客」に乗っ取られ、今は観客の意志を代弁する傀儡
npcs.append({
    'id': 'sukutsu_greed',
    'Author': 'tishi.elin.sukutsu_arena',
    'name_JP': 'グリード',
    'name': 'Greed',
    'aka_JP': '観客の代弁者',
    'aka': 'Voice of the Audience',
    'race': 'wraith',
    'job': 'warrior',
    '_idRenderData': '@chara',
    'tiles': 0,  # human male warrior
    'LV': 150,  # Tier2入口: Lv.100-1,000
    'hostility': 'Enemy',
    'bio': 'm/2003/175/70/possessed',
    'idText': 'sukutsu_greed',
    # 能力: 轟音+混沌ブレス、弱体魔法、宇宙的恐怖、回復
    'mainElement': 'Sound',
    'elements': 'resSound/50,resChaos/30,resMagic/60,featElder/1,featCosmicHorror/1',
    'actCombat': 'breathe_Sound/35,breathe_Chaos/25,SpWeakness/15,ActInsult/30,SpHeal/10',
    # AI: 黒天使のように罵倒する
    'AI_Calm': '4,30',
    'AI_Combat': 'ActInsult',
    'tag': 'boss',
    'quality': 4,
    'chance': 0,
})

# ===== Cランク: 闘技場の鴉（3体ボス）=====
# lore: かつてSランク「屠竜者」を目指した3人の挑戦者
# Bランク以上に進めず、今は番人として闘技場に縛られている

# 12. クロウ（影）- 元盗賊ギルドのエース
npcs.append({
    'id': 'sukutsu_crow_shadow',
    'Author': 'tishi.elin.sukutsu_arena',
    'name_JP': 'クロウ',
    'name': 'Crow',
    'aka_JP': '影の鴉',
    'aka': 'Crow of Shadows',
    'race': 'wraith',
    'job': 'thief',
    '_idRenderData': '@chara',
    'tiles': 2,  # human male thief
    'LV': 350,  # Tier2中盤
    'hostility': 'Enemy',
    'bio': 'm/2004/170/60/silent',
    'idText': 'sukutsu_crow_shadow',
    # 能力: 透明化、闇属性攻撃、地獄攻撃
    'mainElement': 'Darkness',
    'elements': 'invisibility/1,resDarkness/60,resNether/40',
    'actCombat': 'hand_Darkness/40,SpInvisibility/20,hand_Nether/30',
    'tag': 'boss',
    'quality': 4,
    'chance': 0,
})

# 13. レイヴン（刃）- 元戦士ギルドのチャンピオン
npcs.append({
    'id': 'sukutsu_raven_blade',
    'Author': 'tishi.elin.sukutsu_arena',
    'name_JP': 'レイヴン',
    'name': 'Raven',
    'aka_JP': '刃の鴉',
    'aka': 'Raven of Blades',
    'race': 'wraith',
    'job': 'warrior',
    '_idRenderData': '@chara',
    'tiles': 0,  # human male warrior
    'LV': 400,  # Tier2中盤
    'hostility': 'Enemy',
    'bio': 'm/2005/185/90/fierce',
    'idText': 'sukutsu_raven_blade',
    # 能力: 物理特化、追加部位、斬撃/衝撃ブレス
    'mainElement': 'Cut',
    'elements': 'featElder/1,resCut/50,resImpact/40,featBodyParts/2',
    'actCombat': 'breathe_Cut/40,breathe_Impact/30',
    'tag': 'boss',
    'quality': 4,
    'chance': 0,
})

# 14. カラス（毒）- 元錬金術師
npcs.append({
    'id': 'sukutsu_karasu_venom',
    'Author': 'tishi.elin.sukutsu_arena',
    'name_JP': 'カラス',
    'name': 'Karasu',
    'aka_JP': '毒の鴉',
    'aka': 'Crow of Venom',
    'race': 'mutant',
    'job': 'wizard',
    '_idRenderData': '@chara',
    'tiles': 807,  # mutant
    'LV': 450,  # Tier2中盤
    'hostility': 'Enemy',
    'bio': 'f/2006/160/50/cunning',
    'idText': 'sukutsu_karasu_venom',
    # 能力: 毒/酸ブレス、元素の傷跡
    'mainElement': 'Poison',
    'elements': 'resPoison/100,resAcid/60',
    'actCombat': 'breathe_Poison/50,breathe_Acid/30,SpWeakResEle/15',
    'tag': 'boss',
    'quality': 4,
    'chance': 0,
})


# デバッグモードの場合、全キャラのLVを1に設定
if DEBUG_MODE:
    for npc in npcs:
        original_lv = npc.get('LV', 'N/A')
        npc['LV'] = 1
        print(f"  {npc['id']}: LV {original_lv} -> 1")

for npc in npcs:
    rows.append(create_npc_row(npc))

def write_tsv(path, row_data):
    with open(path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f, delimiter='\t', quoting=csv.QUOTE_MINIMAL)
        writer.writerows(row_data)
    print(f'Created TSV: {path}')

write_tsv(OUTPUT_EN_TSV, rows)
write_tsv(OUTPUT_JP_TSV, rows)
