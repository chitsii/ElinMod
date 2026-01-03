"""
アリーナマスター用ドラマスクリプト生成

CWL形式のドラマテーブルを作成
完全カスタムアリーナバトルシステム + 敗北コメント
CWLの動的条件 invoke*/if_flag() を使用
"""

import openpyxl
import os

TOOLS_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(TOOLS_DIR)

OUTPUT_PATH = os.path.join(PROJECT_ROOT, 'LangMod', 'JP', 'Dialog', 'Drama', 'drama_sukutsu_arena_master.xlsx')

# CWL準拠ヘッダー (TinyMita形式)
HEADERS = ['step', 'jump', 'if', 'action', 'param', 'actor', 'version', 'id', 'text_JP', 'text_EN', 'text']

def create_drama():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "master"  # TinyMita形式: キャラクターIDをシート名に

    for col, header in enumerate(HEADERS, 1):
        ws.cell(row=1, column=col, value=header)

    drama_data = [
        # === メインステップ: 動的フラグチェック ===
        # invoke* if_flag() は動的に評価される
        {'step': 'main', 'action': 'invoke*', 'param': 'if_flag(sukutsu_arena_result, ==1)', 'jump': 'victory_comment', 'actor': 'pc'},
        {'action': 'invoke*', 'param': 'if_flag(sukutsu_arena_result, ==2)', 'jump': 'defeat_comment', 'actor': 'pc'},

        # 登録チェック（静的でOK - 一度設定されたら変わらない）
        {'jump': 'registered', 'if': 'hasFlag,sukutsu_gladiator'},

        # === 未登録者への挨拶 ===
        {'id': 'greet1',
         'text_JP': '何の用だ、ひよっこ。見たところ、戦いの「せ」の字も知らなそうだな。',
         'text_EN': "What do you want, greenhorn?"},

        {'action': 'choice', 'jump': 'join_yes', 'id': 'c1',
         'text_JP': '闘士になりたい', 'text_EN': 'I want to become a gladiator'},

        {'action': 'choice', 'jump': 'join_no', 'id': 'c2',
         'text_JP': 'いや、やめておく', 'text_EN': "No, I'll pass"},

        {'action': 'cancel', 'jump': 'end'},

        # ========================================
        # === 勝利時コメント ===
        # ========================================
        # フラグリセット
        {'step': 'victory_comment', 'action': 'invoke*', 'param': 'mod_flag(sukutsu_arena_result, =0)', 'actor': 'pc'},

        {'id': 'win1',
         'text_JP': 'ふん、生きて戻ったか。まあ、最低限の働きはしたようだな。勘違いするなよ、雑魚が。',
         'text_EN': "Humph. You managed to survive? Barely. Don't let it go to your head, worm."},

        {'jump': 'registered'},

        # ========================================
        # === 敗北時コメント（ステージ別）===
        # ========================================
        # フラグリセット
        {'step': 'defeat_comment', 'action': 'invoke*', 'param': 'mod_flag(sukutsu_arena_result, =0)', 'actor': 'pc'},

        # ステージ分岐（動的チェック）
        {'action': 'invoke*', 'param': 'if_flag(sukutsu_arena_failed_stage, >=2)', 'jump': 'defeat_stage2', 'actor': 'pc'},

        # ステージ1敗北
        {'id': 'defeat1',
         'text_JP': 'ハッ、獣にやられたか。情けない。まあ、初心者が負けるのは当然だな。',
         'text_EN': "Ha! Beaten by a beast? Pathetic. Well, beginners always lose."},
        {'id': 'defeat1b',
         'text_JP': '次は足を使え。止まったら死ぬぞ。',
         'text_EN': "Use your legs next time. Stand still and you die."},
        {'jump': 'registered'},

        # ステージ2敗北
        {'step': 'defeat_stage2', 'action': 'invoke*', 'param': 'if_flag(sukutsu_arena_failed_stage, >=3)', 'jump': 'defeat_stage3', 'actor': 'pc'},
        {'id': 'defeat2',
         'text_JP': 'ケンタウロスに蹴り殺されたか？馬並みの速さについていけなかったな。',
         'text_EN': "Trampled by the centaur? Couldn't keep up with that horse-like speed, huh?"},
        {'id': 'defeat2b',
         'text_JP': '逃げるのは悪くない。だが、いつまでも逃げてばかりじゃ勝てないぞ。',
         'text_EN': "Running isn't bad. But you can't win by running forever."},
        {'jump': 'registered'},

        # ステージ3敗北
        {'step': 'defeat_stage3', 'action': 'invoke*', 'param': 'if_flag(sukutsu_arena_failed_stage, >=4)', 'jump': 'defeat_champion', 'actor': 'pc'},
        {'id': 'defeat3',
         'text_JP': 'ミノタウロスか...あれは厄介だったろう。俺も初見では負けたさ。',
         'text_EN': "The minotaur, huh... That one's tough. Even I lost the first time."},
        {'id': 'defeat3b',
         'text_JP': '奴の攻撃を見切れるようになれば、勝機はある。鍛え直して来い。',
         'text_EN': "Learn to read its attacks, and you'll have a chance. Go train more."},
        {'jump': 'registered'},

        # 王者戦敗北
        {'step': 'defeat_champion', 'id': 'defeat4',
         'text_JP': '...グランドマスターに挑んだのか。お前、無謀だな。',
         'text_EN': "...You challenged the Grand Master? You're reckless."},
        {'id': 'defeat4b',
         'text_JP': 'だが、その無謀さは嫌いじゃない。もっと強くなってから、また来い。',
         'text_EN': "But I don't hate that recklessness. Get stronger, then come back."},
        {'jump': 'registered'},

        # ========================================
        # === 登録済み闘士への挨拶 ===
        # ========================================
        {'step': 'registered', 'id': 'greet2',
         'text_JP': 'おう、闘士よ。今日は何の用だ？',
         'text_EN': "Hey, gladiator. What brings you here today?"},

        {'action': 'choice', 'jump': 'battle_prep', 'id': 'c3',
         'text_JP': '戦いに挑む', 'text_EN': 'Challenge the arena'},

        {'action': 'choice', 'jump': 'end', 'id': 'c4',
         'text_JP': 'また今度', 'text_EN': 'Maybe later'},

        {'action': 'cancel', 'jump': 'end'},

        # === 闘士登録：はい ===
        {'step': 'join_yes', 'id': 'join1',
         'text_JP': 'おまえが？ハーッハッハ...まあいい、死にたいなら止めはしない。',
         'text_EN': "You? Ha ha ha... Fine, if you want to die, I won't stop you."},

        {'action': 'setFlag', 'param': 'sukutsu_gladiator,1'},
        {'action': 'setFlag', 'param': 'sukutsu_arena_stage,1'},

        {'id': 'join2',
         'text_JP': 'これでお前も闘士だ。戦いの準備ができたら声をかけろ。',
         'text_EN': "You're a gladiator now. Come talk to me when you're ready to fight."},

        {'action': 'reload', 'jump': 'main'},

        # === 闘士登録：いいえ ===
        {'step': 'join_no', 'id': 'reject1',
         'text_JP': '話は終わりだ。ママのミルクでも飲んでな。',
         'text_EN': "Then we're done here. Go drink your mama's milk."},
        {'jump': 'end'},

        # ========================================
        # === 戦闘準備 - ステージ別アドバイス ===
        # ========================================
        # 動的ステージチェック
        {'step': 'battle_prep', 'action': 'invoke*', 'param': 'if_flag(sukutsu_arena_stage, >=2)', 'jump': 'stage2_prep', 'actor': 'pc'},

        {'id': 'stage1_advice',
         'text_JP': 'お前の最初の相手は「森の狼」だ。素早い攻撃には気をつけろ。',
         'text_EN': "Your first opponent is a Forest Wolf."},
        {'id': 'stage1_tip',
         'text_JP': '武器と防具は整えたか？回復アイテムもあると安心だぞ。',
         'text_EN': "Got your weapons and armor ready?"},

        {'action': 'choice', 'jump': 'battle_start_stage1', 'id': 'c_go1',
         'text_JP': '準備できた、行く！', 'text_EN': "I'm ready, let's go!"},
        {'action': 'choice', 'jump': 'registered', 'id': 'c_cancel1',
         'text_JP': 'もう少し準備してくる', 'text_EN': 'Let me prepare a bit more'},
        {'action': 'cancel', 'jump': 'registered'},

        # ステージ2
        {'step': 'stage2_prep', 'action': 'invoke*', 'param': 'if_flag(sukutsu_arena_stage, >=3)', 'jump': 'stage3_prep', 'actor': 'pc'},
        {'id': 'stage2_advice',
         'text_JP': '次の相手は「ケンタウロス」だ。奴の突進は威力があるぞ。',
         'text_EN': "Your next opponent is a Centaur."},
        {'action': 'choice', 'jump': 'battle_start_stage2', 'id': 'c_go2',
         'text_JP': '準備できた！', 'text_EN': "Ready!"},
        {'action': 'choice', 'jump': 'registered', 'id': 'c_cancel2',
         'text_JP': '待ってくれ', 'text_EN': 'Wait'},
        {'action': 'cancel', 'jump': 'registered'},

        # ステージ3
        {'step': 'stage3_prep', 'action': 'invoke*', 'param': 'if_flag(sukutsu_arena_stage, >=4)', 'jump': 'stage_champion', 'actor': 'pc'},
        {'id': 'stage3_advice',
         'text_JP': 'ここからが本番だ。「ミノタウロス」...奴は俺も手こずった相手だ。',
         'text_EN': "Now the real challenge. The Minotaur..."},
        {'id': 'stage3_tip',
         'text_JP': '力任せに攻めるな。奴の隙を狙え。',
         'text_EN': "Don't just attack blindly."},
        {'action': 'choice', 'jump': 'battle_start_stage3', 'id': 'c_go3',
         'text_JP': '挑む！', 'text_EN': 'Challenge!'},
        {'action': 'choice', 'jump': 'registered', 'id': 'c_cancel3',
         'text_JP': '...もう少し鍛えてくる', 'text_EN': '...I need to train more'},
        {'action': 'cancel', 'jump': 'registered'},

        # 王者戦
        {'step': 'stage_champion', 'id': 'champion_advice',
         'text_JP': 'よくぞここまで来た。最後の相手は...グランドマスターだ。',
         'text_EN': "You've come far. Your final opponent is... the Grand Master."},
        {'id': 'champion_warning',
         'text_JP': '覚悟はいいか？あれは...俺でも勝てるかわからん相手だ。',
         'text_EN': "Are you prepared? Even I... am not sure I could beat that one."},
        {'action': 'choice', 'jump': 'battle_start_champion', 'id': 'c_go_champ',
         'text_JP': '俺は負けない', 'text_EN': "I won't lose"},
        {'action': 'choice', 'jump': 'registered', 'id': 'c_cancel_champ',
         'text_JP': '...考え直す', 'text_EN': '...Let me think about it'},
        {'action': 'cancel', 'jump': 'registered'},

        # ========================================
        # === 戦闘開始 ===
        # ========================================
        {'step': 'battle_start_stage1', 'id': 'sendoff1',
         'text_JP': 'よし、行け！生きて戻ってこい...できればな。',
         'text_EN': "Go! Come back alive... if you can."},
        {'action': 'eval', 'param': 'Elin_SukutsuArena.ArenaManager.StartBattle(tg, 1);'},
        {'action': 'end'},

        {'step': 'battle_start_stage2', 'id': 'sendoff2',
         'text_JP': 'いい度胸だ。お前ならやれる！',
         'text_EN': "Good spirit. You can do this!"},
        {'action': 'eval', 'param': 'Elin_SukutsuArena.ArenaManager.StartBattle(tg, 2);'},
        {'action': 'end'},

        {'step': 'battle_start_stage3', 'id': 'sendoff3',
         'text_JP': '...無茶するなよ。お前はもうただの新人じゃない。',
         'text_EN': "...Don't be reckless."},
        {'action': 'eval', 'param': 'Elin_SukutsuArena.ArenaManager.StartBattle(tg, 3);'},
        {'action': 'end'},

        {'step': 'battle_start_champion', 'id': 'sendoff_champ',
         'text_JP': '...見届けてやる。行って来い、闘士よ。',
         'text_EN': "...I'll be watching. Go, gladiator."},
        {'action': 'eval', 'param': 'Elin_SukutsuArena.ArenaManager.StartBattle(tg, 4);'},
        {'action': 'end'},
    ]

    # Row 6からデータ開始 (TinyMita形式)
    row = 6
    for data in drama_data:
        for col, header in enumerate(HEADERS, 1):
            if header in data:
                ws.cell(row=row, column=col, value=data[header])
        row += 1

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Created: {OUTPUT_PATH}")


if __name__ == "__main__":
    create_drama()
