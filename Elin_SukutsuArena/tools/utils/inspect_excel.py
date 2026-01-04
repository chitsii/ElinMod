import openpyxl
import os

TOOLS_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(TOOLS_DIR)
LANG_EN_DIR = os.path.join(PROJECT_ROOT, 'LangMod', 'EN')

def dump_excel(filename):
    path = os.path.join(LANG_EN_DIR, filename)
    if not os.path.exists(path):
        print(f"File not found: {path}")
        return

    print(f"=== Dumping {filename} ===")
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        for sheetname in wb.sheetnames:
            print(f"--- Sheet: {sheetname} ---")
            ws = wb[sheetname]
            # ヘッダーと最初の数行を表示
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i < 10:  # 最初の10行だけ
                    print(row)
                else:
                    break
    except Exception as e:
        print(f"Error reading {filename}: {e}")

if __name__ == "__main__":
    dump_excel('SourceSukutsu_reference.xlsx')
    # SourceCharaももしあれば
    dump_excel('SourceChara.xlsx')
