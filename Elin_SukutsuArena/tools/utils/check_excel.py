import openpyxl
import os

SAMPLE_PATH = r'c:\Users\tishi\programming\elin_modding\CWL_AddLocation_Example\LangMod\EN\SourceSSS.xlsx'

wb = openpyxl.load_workbook(SAMPLE_PATH)
ws = wb['Chara']

# Header (Row 1)
headers = []
for col in range(1, ws.max_column + 1):
    val = ws.cell(row=1, column=col).value
    if val:  # Only non-None
        headers.append((col, val))

# Data Row 4 - full dump
print("=== Sample NPC Full Data ===")
for col, header in headers:
    val = ws.cell(row=4, column=col).value
    if val is not None and val != "":
        print(f"{header}: {val}")
