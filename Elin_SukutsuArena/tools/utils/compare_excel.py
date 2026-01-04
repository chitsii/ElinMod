import openpyxl

# サンプルと比較
sample_wb = openpyxl.load_workbook(r'c:\Users\tishi\programming\elin_modding\CWL_AddLocation_Example\LangMod\EN\SourceSSS.xlsx')
sample_ws = sample_wb['Zone']

our_wb = openpyxl.load_workbook(r'c:\Users\tishi\programming\elin_modding\ElinMod\Elin_SukutsuArena\LangMod\EN\SourceSukutsu.xlsx')
our_ws = our_wb['Zone']

print('=== サンプル (SourceSSS.xlsx) ===')
print(f'Rows: {sample_ws.max_row}, Cols: {sample_ws.max_column}')
print('Row 1 Headers:')
sample_headers = []
for col in range(1, sample_ws.max_column + 1):
    val = sample_ws.cell(row=1, column=col).value
    sample_headers.append(val)
    print(f'  {col}: {val}')

print('\nRow 4 Data (first 15 cols):')
for col in range(1, min(16, sample_ws.max_column + 1)):
    val = sample_ws.cell(row=4, column=col).value
    print(f'  {col} ({sample_ws.cell(row=1, column=col).value}): {val}')

print('\n=== 私たちの (SourceSukutsu.xlsx) ===')
print(f'Rows: {our_ws.max_row}, Cols: {our_ws.max_column}')
print('Row 1 Headers:')
our_headers = []
for col in range(1, our_ws.max_column + 1):
    val = our_ws.cell(row=1, column=col).value
    our_headers.append(val)
    print(f'  {col}: {val}')

# 差分を確認
print('\n=== 差分 ===')
sample_set = set(sample_headers)
our_set = set(our_headers)
print(f'サンプルにあって私たちにない: {sample_set - our_set}')
print(f'私たちにあってサンプルにない: {our_set - sample_set}')
