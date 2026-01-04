import openpyxl
import os

TOOLS_DIR = os.path.dirname(os.path.abspath(__file__))
# ユーザーのパス環境に合わせて調整
SamplePath = r'c:\Users\tishi\programming\elin_modding\CWL_AddLocation_Example\LangMod\EN\SourceSSS.xlsx'

if os.path.exists(SamplePath):
    wb = openpyxl.load_workbook(SamplePath)
    if 'Chara' in wb.sheetnames:
        ws = wb['Chara']
        headers = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val:
                headers.append(val)
        print("Chara sheet headers:", headers)
    else:
        print("Chara sheet not found.")
else:
    print(f"Sample file not found at {SamplePath}")
